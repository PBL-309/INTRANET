import datetime
import json as pyjson
from flask_mail import Mail, Message
import pandas as pd
from flask import Blueprint, current_app, json, jsonify, render_template, request, redirect, session, url_for, send_from_directory, flash, send_file 
from flask_login import current_user, login_user, logout_user, login_required
import qrcode
from werkzeug.utils import secure_filename
from app import db
from io import BytesIO 
import os
import requests # Necesario para verificar reCAPTCHA v3
import logging
import secrets
import base64
import re
import shutil
import xml.etree.ElementTree as ET
import unicodedata
from functools import wraps
from webauthn import (
    generate_authentication_options,
    generate_registration_options,
    options_to_json,
    verify_authentication_response,
    verify_registration_response,
)
from webauthn.helpers.structs import (
    AuthenticatorSelectionCriteria,
    PublicKeyCredentialDescriptor,
    ResidentKeyRequirement,
    UserVerificationRequirement,
)
from app.models import Aviso, ContactoEmergencia,  Evento, File, Folder, FormularioRespuesta, PortalWeb, Respuesta, User, VacationRequest, Noticia, RegistroCompetencia, EvaluacionDesempeno, AsistenciaFinAnio, PermisosEvaluacion, EntregaUniforme, EntregaGeneralUniforme, MensajeChat, PasskeyCredential, AreaCompra, OrdenCompra, PartidaOrdenCompra, ProveedorCompra, PartidaPresupuestal, FacturaOrdenCompra
from app.forms import LoginForm
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from app import mail
import smtplib
import openpyxl
from decimal import Decimal, InvalidOperation
USUARIOS_RESTRINGIDOS_P5 = {

}
def filtrar_portales_para_usuario(user_id):
    usuario = db.session.get(User, user_id)
    if usuario and usuario.username.lower() == 'admin':
        return PortalWeb.query.order_by(PortalWeb.nombre.asc()).all()
    return PortalWeb.query.filter(
        ~PortalWeb.usuarios_permitidos.any() |
        PortalWeb.usuarios_permitidos.any(User.id == user_id)
    ).order_by(PortalWeb.nombre.asc()).all()
main = Blueprint('main', __name__)
USUARIOS_SIN_FUNCIONES = {'admin', 'admin1'}

@main.route('/entrega_uniformes', methods=['GET', 'POST'])
@login_required
def entrega_uniformes():
    usuarios = User.query.filter(
        ~db.func.lower(User.username).in_(USUARIOS_SIN_FUNCIONES)
    ).order_by(User.nombre.asc(), User.username.asc()).all()
    prendas = ['CHAMARRAS', 'PLAYERA TIPO POLO', 'PLAYERA BLANCA DE VESTIR', 'PANTALONES', 'PANTS', 'BOTAS']

    if request.method == 'POST':
        # Soportar selección por user_id (hidden) o por username de texto
        user_id = request.form.get('user_id', type=int)
        username_input = request.form.get('username', '').strip()
        observaciones = request.form.get('observaciones', '').strip()

        # Normalizar listas de prendas / cantidades para permitir múltiples ítems
        prendas_list = request.form.getlist('prenda[]') or request.form.getlist('prenda')
        cantidades_list = request.form.getlist('cantidad[]') or request.form.getlist('cantidad')

        cantidades_por_prenda = []
        observaciones_por_prenda = {}
        for idx, prenda_catalogo in enumerate(prendas):
            try:
                cantidad = int(request.form.get(f'cantidad_prenda_{idx}', '0') or 0)
            except (TypeError, ValueError):
                cantidad = 0
            if cantidad > 0:
                cantidades_por_prenda.append((prenda_catalogo, cantidad))
                observaciones_por_prenda[prenda_catalogo] = request.form.get(
                    f'observacion_prenda_{idx}', ''
                ).strip()
        if cantidades_por_prenda:
            prendas_list = [item[0] for item in cantidades_por_prenda]
            cantidades_list = [str(item[1]) for item in cantidades_por_prenda]

        # Fallback: si viene un solo par prenda/cantidad como strings
        if not prendas_list and request.form.get('prenda'):
            prendas_list = [request.form.get('prenda')]
        if not cantidades_list and request.form.get('cantidad'):
            cantidades_list = [request.form.get('cantidad')]

        # Validaciones básicas
        if not user_id and not username_input:
            flash('Indica el username del usuario o selecciónalo.', 'warning')
            return redirect(url_for('main.entrega_uniformes'))

        usuario = None
        if user_id:
            usuario = User.query.get(user_id)
        elif username_input:
            usuario = User.query.filter_by(username=username_input).first()

        if not usuario:
            flash('El usuario indicado no existe.', 'danger')
            return redirect(url_for('main.entrega_uniformes'))

        # Crear entregas por cada prenda válida
        created = 0
        first_registro_id = None
        fecha_entrega = datetime.now().replace(microsecond=0)
        try:
            for idx, pr in enumerate(prendas_list):
                pr = (pr or '').strip()
                if not pr or pr not in prendas:
                    continue
                # obtener cantidad correspondiente, si existe
                try:
                    cantidad = int(cantidades_list[idx]) if idx < len(cantidades_list) else 1
                except Exception:
                    cantidad = 1
                if cantidad < 1:
                    continue

                detalle_prenda = observaciones_por_prenda.get(pr, '')
                notas = [detalle_prenda] if detalle_prenda else []
                if observaciones:
                    notas.append(f'Nota general: {observaciones}')
                observacion_registro = ' · '.join(notas)[:250] or None

                entrega = EntregaUniforme(
                    user_id=usuario.id,
                    username=usuario.username,
                    prenda=pr,
                    cantidad=cantidad,
                    observaciones=observacion_registro,
                    fecha_entrega=fecha_entrega
                )
                db.session.add(entrega)
                db.session.flush()
                if first_registro_id is None:
                    first_registro_id = entrega.id
                created += 1

            if created == 0:
                flash('No se encontró ninguna prenda válida para registrar.', 'warning')
                return redirect(url_for('main.entrega_uniformes'))

            db.session.commit()
            flash(f'✅ Se registraron {created} entrega(s) para {usuario.nombre} ({usuario.username}).', 'success')
            return redirect(url_for('main.entrega_uniformes', entrega=first_registro_id))
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception('Error guardando entregas')
            flash(f'Error al guardar las entregas: {e}', 'danger')
            return redirect(url_for('main.entrega_uniformes'))

    all_registros = EntregaUniforme.query.order_by(EntregaUniforme.fecha_entrega.desc()).all()
    filtro = request.args.get('q', '').strip().lower()
    prenda_filtro = request.args.get('prenda', '').strip()
    registros_filtrados = [
        registro for registro in all_registros
        if (not filtro or filtro in registro.username.lower()
            or filtro in ((registro.user.nombre if registro.user else '') or '').lower())
        and (not prenda_filtro or registro.prenda == prenda_filtro)
    ]
    # El historial muestra una sola fila por colaborador. Internamente se
    # conservan las entregas por fecha para mantener la trazabilidad y generar
    # un comprobante independiente de cada visita.
    grupos = {}
    for registro in registros_filtrados:
        if registro.user_id not in grupos:
            grupos[registro.user_id] = {
                'usuario': registro.user,
                'username': registro.username,
                'items': [],
                'entregas': {},
                'total': 0,
                'ultima_fecha': registro.fecha_entrega,
            }
        grupo = grupos[registro.user_id]
        grupo['items'].append(registro)
        grupo['total'] += registro.cantidad
        if registro.fecha_entrega not in grupo['entregas']:
            grupo['entregas'][registro.fecha_entrega] = {
                'id': registro.id,
                'fecha': registro.fecha_entrega,
                'total': 0,
            }
        grupo['entregas'][registro.fecha_entrega]['total'] += registro.cantidad

    for grupo in grupos.values():
        grupo['entregas'] = list(grupo['entregas'].values())
    entregas_recientes = list(grupos.values())[:20]
    registros = registros_filtrados[:20]
    registros_data = []
    for registro in all_registros:
        registros_data.append({
            'id': registro.id,
            'user_id': registro.user_id,
            'username': registro.username,
            'nombre': registro.user.nombre if registro.user else registro.username,
            'prenda': registro.prenda,
            'cantidad': registro.cantidad,
            'observaciones': registro.observaciones or '',
            'fecha_entrega': registro.fecha_entrega.strftime('%Y-%m-%d %H:%M')
        })
    resumen_por_usuario = (
        db.session.query(
            User.username,
            User.nombre,
            db.func.sum(EntregaUniforme.cantidad).label('total')
        )
        .join(EntregaUniforme, EntregaUniforme.user_id == User.id)
        .group_by(User.id)
        .order_by(db.desc('total'))
        .all()
    )
    resumen_por_prenda = (
        db.session.query(
            EntregaUniforme.prenda,
            db.func.sum(EntregaUniforme.cantidad).label('total')
        )
        .group_by(EntregaUniforme.prenda)
        .order_by(db.desc('total'))
        .all()
    )

    total_general = sum(item[1] for item in resumen_por_prenda)
    report_date = datetime.now()

    return render_template(
        'entrega_uniformes.html',
        usuarios=usuarios,
        prendas=prendas,
        registros=registros,
        entregas_recientes=entregas_recientes,
        all_registros=all_registros,
        registros_data=registros_data,
        resumen_por_usuario=resumen_por_usuario,
        resumen_por_prenda=resumen_por_prenda,
        total_general=total_general,
        report_date=report_date,
    )


@main.route('/_search_users')
@login_required
def search_users():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    users = User.query.filter(
        db.or_(User.username.ilike(f"%{q}%"), User.nombre.ilike(f"%{q}%")),
        ~db.func.lower(User.username).in_(USUARIOS_SIN_FUNCIONES)
    )
    users = users.order_by(User.nombre, User.username).limit(12).all()
    result = [{'id': u.id, 'username': u.username, 'nombre': u.nombre} for u in users]
    return jsonify(result)

@main.route('/entrega_uniforme/documento/<int:registro_id>')
@login_required
def entrega_uniforme_documento(registro_id):
    registro = EntregaUniforme.query.get_or_404(registro_id)
    # Un solo reporte por colaborador: incluye todas sus entregas históricas,
    # manteniendo cada artículo y su fecha como un renglón independiente.
    registros = EntregaUniforme.query.filter_by(
        user_id=registro.user_id
    ).order_by(EntregaUniforme.fecha_entrega, EntregaUniforme.id).all()
    return render_template(
        'entrega_uniforme_documento.html',
        registro=registro,
        registros=registros,
        total_piezas=sum(item.cantidad for item in registros),
        total_fechas=len({item.fecha_entrega for item in registros}),
        report_date=datetime.now()
    )


@main.route('/entrega_uniforme/eliminar/<int:registro_id>', methods=['POST'])
@login_required
def eliminar_entrega_uniforme(registro_id):
    registro = EntregaUniforme.query.get_or_404(registro_id)
    colaborador = registro.user.nombre if registro.user else registro.username
    prenda = registro.prenda
    fecha = registro.fecha_entrega.strftime('%d/%m/%Y')
    try:
        db.session.delete(registro)
        db.session.commit()
        flash(f'Se eliminó {prenda.title()} de {colaborador}, entregada el {fecha}.', 'success')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error eliminando entrega de uniforme')
        flash('No fue posible eliminar el registro. Intenta nuevamente.', 'danger')
    return redirect(url_for('main.entrega_uniformes'))


@main.route('/entrega_uniformes_general', methods=['GET', 'POST'])
@login_required
def entrega_uniformes_general():
    prendas = ['CHAQUETA', 'PLAYERA TIPO POLO', 'PLAYERA BLANCA DE VESTIR', 'PANTALÓN', 'BOTAS']

    if request.method == 'POST':
        receptor_id = request.form.get('receptor_id', type=int)
        receptor = db.session.get(User, receptor_id) if receptor_id else None
        if not receptor:
            flash('Selecciona a la persona que recibe los uniformes.', 'warning')
            return redirect(url_for('main.entrega_uniformes_general'))

        items = []
        nota_general = request.form.get('nota_general', '').strip()
        for idx, prenda in enumerate(prendas):
            try:
                cantidad = int(request.form.get(f'cantidad_{idx}', '0') or 0)
            except (TypeError, ValueError):
                cantidad = 0
            if cantidad > 0:
                detalle = request.form.get(f'detalle_{idx}', '').strip()
                textos = [detalle] if detalle else []
                if nota_general:
                    textos.append(f'Nota general: {nota_general}')
                items.append((prenda, cantidad, ' · '.join(textos)[:250] or None))

        if not items:
            flash('Agrega al menos una prenda para registrar la entrega.', 'warning')
            return redirect(url_for('main.entrega_uniformes_general'))

        fecha = datetime.now().replace(microsecond=0)
        try:
            for prenda, cantidad, detalle in items:
                db.session.add(EntregaGeneralUniforme(
                    receptor_id=receptor.id,
                    entregado_por_id=current_user.id,
                    prenda=prenda,
                    cantidad=cantidad,
                    detalle=detalle,
                    fecha_entrega=fecha,
                ))
            db.session.commit()
            flash(f'Entrega registrada para {receptor.nombre}.', 'success')
            return redirect(url_for('main.entrega_uniformes_general', receptor=receptor.id))
        except Exception:
            db.session.rollback()
            current_app.logger.exception('Error registrando entrega general de uniformes')
            flash('No fue posible guardar la entrega. Intenta nuevamente.', 'danger')
            return redirect(url_for('main.entrega_uniformes_general'))

    usuarios = User.query.filter(
        ~db.func.lower(User.username).in_(USUARIOS_SIN_FUNCIONES)
    ).order_by(User.nombre, User.username).all()
    registros = EntregaGeneralUniforme.query.order_by(EntregaGeneralUniforme.fecha_entrega.desc()).all()
    por_receptor = {}
    for registro in registros:
        if registro.receptor_id not in por_receptor:
            por_receptor[registro.receptor_id] = {
                'receptor': registro.receptor,
                'items': [],
                'fechas': set(),
                'total': 0,
                'ultima_fecha': registro.fecha_entrega,
                'ultimo_entregador': registro.entregado_por,
            }
        grupo = por_receptor[registro.receptor_id]
        grupo['items'].append(registro)
        grupo['fechas'].add(registro.fecha_entrega)
        grupo['total'] += registro.cantidad

    atendidos_ids = set(por_receptor)
    pendientes = [usuario for usuario in usuarios if usuario.id not in atendidos_ids]
    turnos_con_entregas = sorted({
        (grupo['receptor'].turno or '').strip()
        for grupo in por_receptor.values()
        if (grupo['receptor'].turno or '').strip()
    })
    return render_template(
        'entrega_uniformes_general.html',
        prendas=prendas,
        usuarios=usuarios,
        entregas=list(por_receptor.values()),
        pendientes=pendientes,
        atendidos=len(atendidos_ids),
        total_personal=len(usuarios),
        total_piezas=sum(registro.cantidad for registro in registros),
        turnos_con_entregas=turnos_con_entregas,
    )


@main.route('/entrega_uniformes_general/reporte/<int:receptor_id>')
@login_required
def reporte_entrega_uniformes_general(receptor_id):
    receptor = db.session.get(User, receptor_id)
    if not receptor:
        return ('', 404)
    registros = EntregaGeneralUniforme.query.filter_by(receptor_id=receptor_id).order_by(
        EntregaGeneralUniforme.fecha_entrega, EntregaGeneralUniforme.id
    ).all()
    if not registros:
        return ('', 404)
    return render_template(
        'entrega_uniformes_general_reporte.html',
        receptor=receptor,
        registros=registros,
        total_piezas=sum(item.cantidad for item in registros),
        total_fechas=len({item.fecha_entrega for item in registros}),
        report_date=datetime.now(),
    )


def _numero_nomina_orden(username):
    """Ordena primero las nóminas numéricas por su valor real."""
    texto = str(username or '').strip()
    return (0, int(texto), texto) if texto.isdigit() else (1, 0, texto.casefold())


@main.route('/entrega_uniformes_general/reportes/turno/<path:turno>')
@login_required
def reportes_entrega_uniformes_por_turno(turno):
    turno_normalizado = (turno or '').strip()
    if not turno_normalizado:
        return ('', 404)

    registros = (
        EntregaGeneralUniforme.query
        .join(User, User.id == EntregaGeneralUniforme.receptor_id)
        .filter(db.func.lower(db.func.trim(User.turno)) == turno_normalizado.lower())
        .filter(~db.func.lower(User.username).in_(USUARIOS_SIN_FUNCIONES))
        .order_by(EntregaGeneralUniforme.fecha_entrega, EntregaGeneralUniforme.id)
        .all()
    )
    por_receptor = {}
    for registro in registros:
        por_receptor.setdefault(registro.receptor_id, {
            'receptor': registro.receptor,
            'registros': [],
        })['registros'].append(registro)
    reportes = sorted(
        por_receptor.values(),
        key=lambda reporte: _numero_nomina_orden(reporte['receptor'].username),
    )
    if not reportes:
        return ('', 404)
    for reporte in reportes:
        items = reporte['registros']
        reporte['total_piezas'] = sum(item.cantidad for item in items)
        reporte['total_fechas'] = len({item.fecha_entrega.date() for item in items})

    return render_template(
        'entrega_uniformes_general_lote.html',
        turno=turno_normalizado,
        reportes=reportes,
        report_date=datetime.now(),
    )


@main.route('/entrega_uniformes_general/eliminar/<int:registro_id>', methods=['POST'])
@login_required
def eliminar_entrega_uniformes_general(registro_id):
    registro = EntregaGeneralUniforme.query.get_or_404(registro_id)
    nombre = registro.receptor.nombre
    try:
        db.session.delete(registro)
        db.session.commit()
        flash(f'Registro de {registro.prenda.title()} eliminado para {nombre}.', 'success')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error eliminando entrega general de uniformes')
        flash('No fue posible eliminar el registro.', 'danger')
    return redirect(url_for('main.entrega_uniformes_general'))


@main.route('/api/entrega_uniformes_general/persona/<int:receptor_id>')
@login_required
def articulos_entrega_general(receptor_id):
    receptor = db.session.get(User, receptor_id)
    if not receptor:
        return jsonify({'error': 'Persona no encontrada.'}), 404
    registros = EntregaGeneralUniforme.query.filter_by(receptor_id=receptor_id).order_by(
        EntregaGeneralUniforme.fecha_entrega.desc(), EntregaGeneralUniforme.id.desc()
    ).all()
    return jsonify({
        'receptor': {'id': receptor.id, 'nombre': receptor.nombre, 'username': receptor.username},
        'articulos': [{
            'id': item.id,
            'prenda': item.prenda,
            'cantidad': item.cantidad,
            'detalle': item.detalle or '',
            'fecha': item.fecha_entrega.strftime('%Y-%m-%dT%H:%M'),
        } for item in registros]
    })


@main.route('/api/entrega_uniformes_general/persona/<int:receptor_id>', methods=['PUT'])
@login_required
def actualizar_articulos_entrega_general(receptor_id):
    prendas_validas = {'CHAQUETA', 'PLAYERA TIPO POLO', 'PLAYERA BLANCA DE VESTIR', 'PANTALÓN', 'BOTAS'}
    data = request.get_json(silent=True) or {}
    cambios = data.get('articulos', [])
    if not isinstance(cambios, list) or not cambios:
        return jsonify({'error': 'No hay cambios para guardar.'}), 400
    registros = {
        item.id: item for item in EntregaGeneralUniforme.query.filter_by(receptor_id=receptor_id).all()
    }
    try:
        actualizados = eliminados = 0
        for cambio in cambios:
            try:
                registro_id = int(cambio.get('id'))
            except (TypeError, ValueError):
                continue
            registro = registros.get(registro_id)
            if not registro:
                continue
            if cambio.get('eliminar'):
                db.session.delete(registro)
                eliminados += 1
                continue
            prenda = str(cambio.get('prenda', '')).strip().upper()
            cantidad = int(cambio.get('cantidad', 0))
            if prenda not in prendas_validas or cantidad < 1 or cantidad > 99:
                raise ValueError('Revisa el artículo y su cantidad.')
            registro.prenda = prenda
            registro.cantidad = cantidad
            registro.detalle = str(cambio.get('detalle', '')).strip()[:250] or None
            fecha_texto = str(cambio.get('fecha', '')).strip()
            if fecha_texto:
                registro.fecha_entrega = datetime.fromisoformat(fecha_texto)
            actualizados += 1
        db.session.commit()
        return jsonify({'ok': True, 'actualizados': actualizados, 'eliminados': eliminados})
    except (ValueError, TypeError):
        db.session.rollback()
        return jsonify({'error': 'Revisa cantidades, artículos y fechas.'}), 400
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error actualizando artículos de entrega general')
        return jsonify({'error': 'No fue posible actualizar los artículos.'}), 500


def _chat_user_data(user):
    foto_filename = f'uploads/{user.username}.jpg'
    foto_path = os.path.join(current_app.root_path, 'static', foto_filename)
    if not os.path.exists(foto_path):
        foto_filename = 'uploads/default.png'
    conectado = bool(user.ultima_actividad and user.ultima_actividad >= datetime.now() - timedelta(seconds=75))
    return {
        'id': user.id,
        'username': user.username,
        'nombre': user.nombre,
        'iniciales': ''.join(parte[0] for parte in user.nombre.split()[:2]).upper(),
        'foto_url': url_for('static', filename=foto_filename),
        'conectado': conectado,
    }


@main.route('/api/chat/usuarios')
@login_required
def chat_usuarios():
    q = request.args.get('q', '').strip()
    query = User.query.filter(
        User.id != current_user.id,
        ~db.func.lower(User.username).in_(USUARIOS_SIN_FUNCIONES),
    )
    if q:
        query = query.filter(db.or_(User.nombre.ilike(f'%{q}%'), User.username.ilike(f'%{q}%')))
    users = query.order_by(User.nombre).limit(30).all()
    return jsonify([_chat_user_data(user) for user in users])


@main.route('/api/chat/presencia')
@login_required
def chat_presencia():
    current_user.ultima_actividad = datetime.now()
    db.session.commit()
    return jsonify({'ok': True})


@main.route('/api/chat/conversaciones')
@login_required
def chat_conversaciones():
    mensajes = MensajeChat.query.filter(
        db.or_(
            MensajeChat.remitente_id == current_user.id,
            MensajeChat.destinatario_id == current_user.id,
        )
    ).order_by(MensajeChat.fecha_envio.desc()).limit(500).all()

    conversaciones = {}
    for mensaje in mensajes:
        contacto = mensaje.destinatario if mensaje.remitente_id == current_user.id else mensaje.remitente
        if contacto.id not in conversaciones:
            conversaciones[contacto.id] = {
                **_chat_user_data(contacto),
                'ultimo_mensaje': mensaje.contenido,
                'ultimo_mensaje_id': mensaje.id,
                'ultimo_mensaje_propio': mensaje.remitente_id == current_user.id,
                'fecha': mensaje.fecha_envio.strftime('%d/%m %H:%M'),
                'no_leidos': 0,
            }
        if mensaje.destinatario_id == current_user.id and mensaje.leido_en is None:
            conversaciones[contacto.id]['no_leidos'] += 1

    total_no_leidos = sum(item['no_leidos'] for item in conversaciones.values())
    return jsonify({'conversaciones': list(conversaciones.values()), 'no_leidos': total_no_leidos})


@main.route('/api/chat/mensajes/<int:contacto_id>')
@login_required
def chat_mensajes(contacto_id):
    contacto = db.session.get(User, contacto_id)
    if not contacto or contacto.id == current_user.id:
        return jsonify({'error': 'Usuario no válido.'}), 404

    MensajeChat.query.filter_by(
        remitente_id=contacto.id,
        destinatario_id=current_user.id,
        leido_en=None,
    ).update({'leido_en': datetime.now()}, synchronize_session=False)
    db.session.commit()

    mensajes = MensajeChat.query.filter(
        db.or_(
            db.and_(MensajeChat.remitente_id == current_user.id, MensajeChat.destinatario_id == contacto.id),
            db.and_(MensajeChat.remitente_id == contacto.id, MensajeChat.destinatario_id == current_user.id),
        )
    ).order_by(MensajeChat.fecha_envio.desc()).limit(100).all()
    mensajes.reverse()
    return jsonify({
        'contacto': _chat_user_data(contacto),
        'mensajes': [{
            'id': mensaje.id,
            'contenido': mensaje.contenido,
            'propio': mensaje.remitente_id == current_user.id,
            'fecha': mensaje.fecha_envio.strftime('%d/%m/%Y %H:%M'),
            'leido': mensaje.leido_en is not None,
            'editado': mensaje.editado_en is not None,
        } for mensaje in mensajes],
    })


@main.route('/api/chat/mensajes', methods=['POST'])
@login_required
def enviar_mensaje_chat():
    data = request.get_json(silent=True) or {}
    destinatario_id = data.get('destinatario_id')
    contenido = str(data.get('contenido', '')).strip()
    destinatario = db.session.get(User, destinatario_id) if destinatario_id else None
    if not destinatario or destinatario.id == current_user.id:
        return jsonify({'error': 'Selecciona un destinatario válido.'}), 400
    if destinatario.username.lower() in USUARIOS_SIN_FUNCIONES:
        return jsonify({'error': 'Este usuario no está disponible en el chat.'}), 400
    if not contenido:
        return jsonify({'error': 'Escribe un mensaje.'}), 400
    if len(contenido) > 1000:
        return jsonify({'error': 'El mensaje es demasiado largo.'}), 400
    try:
        mensaje = MensajeChat(
            remitente_id=current_user.id,
            destinatario_id=destinatario.id,
            contenido=contenido,
            fecha_envio=datetime.now(),
        )
        db.session.add(mensaje)
        db.session.commit()
        return jsonify({'ok': True, 'id': mensaje.id})
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error enviando mensaje de chat')
        return jsonify({'error': 'No fue posible enviar el mensaje.'}), 500


@main.route('/api/chat/mensajes/<int:mensaje_id>', methods=['PUT', 'DELETE'])
@login_required
def modificar_mensaje_chat(mensaje_id):
    mensaje = MensajeChat.query.filter_by(id=mensaje_id, remitente_id=current_user.id).first()
    if not mensaje:
        return jsonify({'error': 'Mensaje no encontrado o sin permiso.'}), 404
    try:
        if request.method == 'DELETE':
            db.session.delete(mensaje)
            db.session.commit()
            return jsonify({'ok': True})
        contenido = str((request.get_json(silent=True) or {}).get('contenido', '')).strip()
        if not contenido:
            return jsonify({'error': 'El mensaje no puede quedar vacío.'}), 400
        if len(contenido) > 1000:
            return jsonify({'error': 'El mensaje es demasiado largo.'}), 400
        mensaje.contenido = contenido
        mensaje.editado_en = datetime.now()
        db.session.commit()
        return jsonify({'ok': True, 'contenido': mensaje.contenido})
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error modificando mensaje de chat')
        return jsonify({'error': 'No fue posible modificar el mensaje.'}), 500

@main.route('/consultar_evaluaciones')
@login_required
def consultar_evaluaciones():
    allowed_ids = [304]  
    if current_user.id not in allowed_ids:
        return """
        <script>
            alert("No tienes los permisos necesarios para ver esta información.");
            window.location.href = "/";
        </script>
        """
    preguntas = [
    "Conozco mis funciones y mis responsabilidades",
    "Al realizar mi trabajo, siempre doy resultados positivos",
    "Las actividades que realizo me permiten desarrollar mis habilidades",
    "Tengo conocimiento de las políticas y reglamentos de la institución",
    "Puedo agilizar el trabajo actuando antes de que me lo pidan",
    "Ofrezco ayuda sin necesidad de que lo soliciten",
    "Me mantiene informado sobre asuntos que afectan a mi trabajo",
    "Pone el ejemplo en la forma en la que debe ser nuestro desempeño",
    "Soluciona los problemas de manera eficaz",
    "Encomienda las actividades de trabajo de manera igualitaria entre los compañeros",
    "Trata de igual manera a todo el personal y se comunica con respeto",
    "Identifica áreas de mejora optimizando el desempeño de nuestras funciones",
    "Siempre hay colaboración con mis compañeros para el buen desempeño del trabajo",
    "Me siento parte de un equipo de trabajo",
    "Es fácil expresar las opiniones ante compañeros y jefe inmediato",
    "Se generan planes de trabajo positivos que motivan al personal",
    "Se tiene un ambiente sano y cordial entre los compañeros de las otras estaciones",
    "Existe colaboración entre los compañeros de toda la institución para la realización de las tareas",
    "El nombre y prestigio de la institución me hacen sentir orgulloso de pertenecer a ella",
    "Conozco el trabajo que realizan las diferentes áreas de la institución",
    "Las estaciones cuentan con una instalación que permite sentirme a gusto",
    "Recibo capacitaciones constantes que permiten actualizar mis conocimientos sobre el trabajo",
    "Doy un buen uso y cuidado a las herramientas y equipo otorgado por la institución",
    "Me gusta la estación a la que pertenezco",
    "Descubro los intereses de mis compañeros e intento llevarlos a una meta en común",
    "Intento ser honesto y transparente cuando hablo",
    "Me considero una persona que explica con detalle cuando capacito a mis compañeros",
    "Conozco a mis compañeros y tengo muy buena relación con ellos",
    "Intento ofrecer todo mi apoyo cuando un compañero pasa por un momento de dificultad personal",
    "Escucho, respeto y reflexiono sobre las opiniones de los demás antes de tomar decisiones",
    "Me gusta dirigir las reuniones de trabajo para que todos participen",
    "Supervisar de cerca el equipo es buena opción para asegurarme que las cosas salgan bien",
    "Me cuesta trabajo admitir mis errores y suelo culpar a los demás",
    "Es bueno que nos asignen tareas sin decidir si está bien o no",
    "No acepto mis fracasos personales ni la opinión de que estoy equivocado en mi forma de ver las cosas",
    "Exijo superar las metas que consigue mi equipo con el fin de seguir mejorando",
    "Tengo una comunicación abierta con mis compañeros/personal a cargo y conozco lo que esperan de mí",
    "Los errores y fracasos de los demás los utilizo como motivación para superarlos",
    "Me gusta pensar de manera positiva para motivar al personal que me rodea y cumplir con los objetivos que se esperan en el trabajo",
    "Me da miedo tomar el control y prefiero que otros se encarguen",
    "Busco opiniones de mis compañeros para lograr ideas y proyectos nuevos",
    "Conseguir resultados positivos es mi prioridad, la tensión es el precio a pagar por el éxito",
    "Cuando hay problemas en mi equipo, me pongo muy nervioso y no sé cómo controlarme",
    "Soy una persona protectora con los compañeros que piensan igual que yo y obedecen, respeto a los que me contradices pero me cuesta mucho confiar en los que mienten"
]

    respuesta_texto = {
        "1": "Totalmente en desacuerdo",
        "2": "En desacuerdo",
        "3": "Indiferente",
        "4": "De acuerdo",
        "5": "Totalmente de acuerdo"
    }

    orden = request.args.get("orden", "desc")  
    seleccionado_id = request.args.get("id")  

    if seleccionado_id:  
        evaluaciones = EvaluacionDesempeno.query.filter_by(id=seleccionado_id).all()
    else:
        evaluaciones = EvaluacionDesempeno.query.all()

    evaluaciones_data = []
    for evaluacion in evaluaciones:
        respuestas = evaluacion.respuestas
        total = 0
        count = 0
        promedios_rubro = {
            "Desempeño": 0,
            "Jefe Inmediato": 0,
            "Apoyo y Convivencia": 0,
            "Pertenencia": 0,
            "Clasificación": 0
        }
        for rubro, preguntas in respuestas.items():
            valores = [int(v) for v in preguntas.values() if v]
            if valores:
                promedio_rubro = round(sum(valores) / len(valores), 2)
            else:
                promedio_rubro = 0

            if rubro == "desempeno":
                promedios_rubro["Desempeño"] = promedio_rubro
            elif rubro == "jefe_inmediato":
                promedios_rubro["Jefe Inmediato"] = promedio_rubro
            elif rubro == "apoyo_convivencia":
                promedios_rubro["Apoyo y Convivencia"] = promedio_rubro
            elif rubro == "pertenencia":
                promedios_rubro["Pertenencia"] = promedio_rubro
            elif rubro == "clasificacion":
                promedios_rubro["Clasificación"] = promedio_rubro

            total += sum(valores)
            count += len(valores)

        promedio = round(total / count, 2) if count > 0 else 0

        evaluaciones_data.append({
            "id": evaluacion.id,
            "nombre": evaluacion.nombre,
            "respuestas": respuestas,
            "promedio": promedio,
            "promedios_rubro": promedios_rubro
        })
    reverse = (orden == "desc")
    evaluaciones_data.sort(key=lambda x: x["promedio"], reverse=reverse)
    promedios = {
        "Desempeño": [],
        "Jefe Inmediato": [],
        "Apoyo y Convivencia": [],
        "Pertenencia": [],
        "Clasificación": []
    }
    conteo = {k: 0 for k in promedios}

    for e in evaluaciones:
        for rubro, preguntas in e.respuestas.items():
            valores = [int(v) for v in preguntas.values() if v]
            if valores:
                promedio_rubro = round(sum(valores) / len(valores), 2)
                if rubro == "desempeno":
                    promedios["Desempeño"].append(promedio_rubro)
                    conteo["Desempeño"] += 1
                elif rubro == "jefe_inmediato":
                    promedios["Jefe Inmediato"].append(promedio_rubro)
                    conteo["Jefe Inmediato"] += 1
                elif rubro == "apoyo_convivencia":
                    promedios["Apoyo y Convivencia"].append(promedio_rubro)
                    conteo["Apoyo y Convivencia"] += 1
                elif rubro == "pertenencia":
                    promedios["Pertenencia"].append(promedio_rubro)
                    conteo["Pertenencia"] += 1
                elif rubro == "clasificacion":
                    promedios["Clasificación"].append(promedio_rubro)
                    conteo["Clasificación"] += 1

    # promedio final por rubro
    promedios_finales = []
    for rubro, valores in promedios.items():
        if valores:
            promedios_finales.append(round(sum(valores) / len(valores), 2))
        else:
            promedios_finales.append(0)

    return render_template(
        'consultar_evaluaciones.html',
        evaluaciones=evaluaciones_data,
        promedios=promedios_finales,
        respuesta_texto=respuesta_texto,
        orden=orden,
        seleccionado_id=seleccionado_id  
    )

@main.route('/descargar_asistencia_excel_formato')
@login_required
def descargar_asistencia_excel_formato():
    asistentes_data = db.session.query(
        AsistenciaFinAnio.nombre_usuario,
        User.username,
        User.turno,  
        AsistenciaFinAnio.lleva_acompanante,
        AsistenciaFinAnio.fecha_registro
    ).join(User, AsistenciaFinAnio.user_id == User.id) \
     .filter(AsistenciaFinAnio.asistencia == 'sí') \
     .order_by(AsistenciaFinAnio.nombre_usuario) \
     .all()
    data = []
    for nombre_usuario, username, turno, lleva_acompanante, fecha_registro in asistentes_data:
        lleva_acompanante_str = "Sí" if lleva_acompanante == 'sí' else "No"
        data.append({
            'Nombre del Empleado': nombre_usuario,
            'Username': username,
            'Turno': turno if turno else 'N/A', 
            'Asistencia': 'Sí',
            'Lleva Acompañante': lleva_acompanante_str,
            'Fecha de Registro': fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
        })
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Asistentes', index=False)
        worksheet = writer.sheets['Asistentes']
        header_font = openpyxl.styles.Font(bold=True)
        side_border = openpyxl.styles.Side(style='thin')
        full_border = openpyxl.styles.Border(left=side_border, 
                                             right=side_border, 
                                             top=side_border, 
                                             bottom=side_border)
        for col_num, value in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = value
            cell.font = header_font
            cell.border = full_border
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    cell.border = full_border
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
    output.seek(0)
    filename = "Lista_Asistencia_Fin_Anio.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )        


@main.route('/registro_fin_anio12w', methods=['GET', 'POST'])
@login_required
def submit_fin():
    if request.method == 'POST':
        asistencia = request.form.get('asistencia')
        lleva_acompanante = request.form.get('lleva_acompanante')
        user_id = current_user.id
        nombre_del_usuario = current_user.nombre 
        
        if not asistencia or not lleva_acompanante:
            flash('Por favor, responde a todas las preguntas de asistencia.', 'danger')
            return redirect(url_for('main.submit_fin'))
            
        registro_existente = AsistenciaFinAnio.query.filter_by(user_id=user_id).first()
        
        try:
            if registro_existente:
                registro_existente.asistencia = asistencia
                registro_existente.lleva_acompanante = lleva_acompanante
                registro_existente.nombre_usuario = nombre_del_usuario 
                flash('✅ Tu registro de asistencia ha sido **actualizado** exitosamente.', 'success')
            else:
                nuevo_registro = AsistenciaFinAnio(
                    user_id=user_id,
                    nombre_usuario=nombre_del_usuario, 
                    asistencia=asistencia,
                    lleva_acompanante=lleva_acompanante
                )
                db.session.add(nuevo_registro)
                flash('💾 Tu registro ha sido **guardado** exitosamente. ¡Gracias!', 'success')
            db.session.commit()
            if asistencia == 'sí': 

                return redirect(url_for('main.confirmacion_exitosa'))

            else:
                return redirect(url_for('main.cancelacion'))
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Ocurrió un error al guardar tu respuesta: {e}', 'danger')
            return redirect(url_for('main.submit_fin'))
    registro = AsistenciaFinAnio.query.filter_by(user_id=current_user.id).first()
    return render_template('confirmation.html', registro=registro)



@main.route('/dashboard')
@login_required
def dashboard():
    uploads_path = current_app.config['UPLOAD_FOLDER']
    static_uploads_path = os.path.join(current_app.root_path, 'static', 'uploads')
    username = current_user.username
    user_image_path = os.path.join(static_uploads_path, f"{username}.jpg")

    if os.path.exists(user_image_path):
        user_image = f"uploads/{username}.jpg"
    else:
        user_image = "uploads/default.png"
    files_in_db = File.query.all()
    visible_files = []
    for file in files_in_db:
        filepath = os.path.join(uploads_path, file.filename)
        if os.path.exists(filepath):
            visible_files.append(file)
    avisos = Aviso.query.order_by(Aviso.fecha_creacion.desc()).all()
    eventos = Evento.query.order_by(Evento.fecha.asc()).all()
    noticias = Noticia.query.order_by(Noticia.orden.asc()).all()
    portales = filtrar_portales_para_usuario(current_user.id)
    usuarios_portales = []
    if current_user.username.lower() == 'admin':
        usuarios_portales = User.query.filter(
            ~db.func.lower(User.username).in_(USUARIOS_SIN_FUNCIONES)
        ).order_by(User.nombre.asc(), User.username.asc()).all()

    return render_template(
        'dashboard.html',
        files=visible_files,  
        avisos=avisos,
        eventos=eventos,
        noticias=noticias,
        portales=portales,
        usuarios_portales=usuarios_portales,
        user_image=user_image,
        puede_compras=_codigo_area_compra_usuario(current_user) is not None,
    )


USUARIOS_AREAS_COMPRA = {
    'CRISTIAN ADAN RODRIGUEZ CARDENAS': 'SP', 'ANA VALERIA GARCIA MORENO': 'SP',
    'EVA GOMEZ HERRERA': 'SP', 'OMAR FERNANDO HERNANDEZ SOLIS': 'PV',
    'MARIA CLAUDIA YEPEZ GUTIERREZ': 'CONT', 'VICTOR HUGO BECERRA MORENO': 'CONT',
    'MIGUEL ANGEL DE ANDA AGUINAGA': 'CONT', 'MARIA DEL CARMEN MORENO GUTIERREZ': 'CONT',
    'ISMAEL MURILLO BOLIVAR': 'CH', 'SALMA KARIME ORTIZ GARNICA': 'CH',
    'KARLA VANESSA RIOS AVINA': 'CH',
}


def _texto_sin_acentos(valor):
    return ''.join(c for c in unicodedata.normalize('NFD', valor or '') if unicodedata.category(c) != 'Mn').upper().strip()


def _codigo_area_compra_usuario(usuario):
    if not usuario or not usuario.is_authenticated:
        return None
    if (usuario.username or '').lower() in {'admin', 'admin1'}:
        return '*'
    return USUARIOS_AREAS_COMPRA.get(_texto_sin_acentos(usuario.nombre))


def _areas_compra_usuario():
    codigo = _codigo_area_compra_usuario(current_user)
    consulta = AreaCompra.query.filter_by(activa=True)
    if codigo == '*':
        return consulta.order_by(AreaCompra.nombre).all()
    return consulta.filter_by(codigo=codigo).all() if codigo else []


def acceso_compras_requerido(funcion):
    @wraps(funcion)
    def protegida(*args, **kwargs):
        if _codigo_area_compra_usuario(current_user) is None:
            flash('Tu usuario no tiene acceso al módulo de órdenes de compra.', 'warning')
            return redirect(url_for('main.dashboard'))
        return funcion(*args, **kwargs)
    return protegida


def _areas_compra_iniciales():
    """Crea las series autorizadas sin reducir contadores ya utilizados."""
    series = {'SP': ('SOPORTE TÉCNICO', 81), 'PV': ('PREVENCIÓN', 576), 'CONT': ('CONTABILIDAD', 0), 'CH': ('CAPITAL HUMANO', 0)}
    cambio = False
    for codigo, (nombre, ultimo) in series.items():
        area = AreaCompra.query.filter_by(codigo=codigo).first()
        if not area:
            db.session.add(AreaCompra(nombre=nombre, codigo=codigo, ultimo_consecutivo=ultimo))
            cambio = True
        elif not area.activa:
            area.activa = True
            cambio = True
    if cambio:
        db.session.commit()


def _decimal_factura(valor, default='0'):
    texto = str(valor or '').replace('$', '').replace(',', '').strip()
    try:
        return Decimal(texto)
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _extraer_factura_xml(ruta):
    raiz = ET.parse(ruta).getroot()
    local = lambda etiqueta: etiqueta.rsplit('}', 1)[-1]
    if local(raiz.tag).lower() != 'comprobante':
        raise ValueError('El XML no parece ser un CFDI válido.')
    emisor = next((n for n in raiz.iter() if local(n.tag) == 'Emisor'), None)
    conceptos = [n for n in raiz.iter() if local(n.tag) == 'Concepto']
    traslados = [n for n in raiz.iter() if local(n.tag) == 'Traslado' and n.attrib.get('Impuesto') == '002']
    fecha_texto = raiz.attrib.get('Fecha', '')[:10]
    try:
        fecha = datetime.strptime(fecha_texto, '%Y-%m-%d').date().isoformat()
    except ValueError:
        fecha = datetime.now().date().isoformat()
    items = []
    for concepto in conceptos:
        descripcion = concepto.attrib.get('Descripcion', '').strip()
        if descripcion:
            items.append({
                'cantidad': str(_decimal_factura(concepto.attrib.get('Cantidad'), '1')),
                'descripcion': descripcion,
                'precio_unitario': str(_decimal_factura(concepto.attrib.get('ValorUnitario'))),
                'importe': str(_decimal_factura(concepto.attrib.get('Importe'))),
            })
    iva = sum((_decimal_factura(n.attrib.get('Importe')) for n in traslados), Decimal('0'))
    return {
        'origen': 'XML CFDI',
        'proveedor': (emisor.attrib.get('Nombre', '') if emisor is not None else '').strip().upper(),
        'rfc': (emisor.attrib.get('Rfc', '') if emisor is not None else '').strip().upper(),
        'fecha': fecha,
        'subtotal': str(_decimal_factura(raiz.attrib.get('SubTotal'))),
        'iva': str(iva),
        'total': str(_decimal_factura(raiz.attrib.get('Total'))),
        'moneda': raiz.attrib.get('Moneda', 'MXN'),
        'domicilio': '',
        'telefono': '',
        'conceptos': items,
    }


def _buscar_importe_pdf(texto, etiqueta):
    patrones = [
        rf'(?im)^\s*{etiqueta}\s*[:$]?\s*\$?\s*([\d,]+\.\d{{2}})',
        rf'(?im){etiqueta}[^\d]{{0,20}}\$?\s*([\d,]+\.\d{{2}})',
    ]
    for patron in patrones:
        coincidencias = re.findall(patron, texto)
        if coincidencias:
            return str(_decimal_factura(coincidencias[-1]))
    return '0'


def _extraer_conceptos_pdf(lineas):
    conceptos = []
    indice = 0
    numero = re.compile(r'^\$?\s*([\d,]+\.\d{2})\s*$')
    cantidad = re.compile(r'^\d+(?:\.\d+)?$')
    while indice < len(lineas) - 2:
        if not cantidad.fullmatch(lineas[indice]) or not re.search(r'(?i)(pieza|servicio|unidad|H87|ACT|E48)', lineas[indice + 1]):
            indice += 1
            continue
        inicio = indice
        cantidad_texto = lineas[indice]
        indice += 2
        descripcion = []
        while indice < len(lineas) and not numero.fullmatch(lineas[indice]) and not re.match(r'(?i)^subtotal', lineas[indice]):
            descripcion.append(lineas[indice])
            indice += 1
        if indice >= len(lineas) or not numero.fullmatch(lineas[indice]):
            indice = inicio + 1
            continue
        precio = str(_decimal_factura(numero.fullmatch(lineas[indice]).group(1)))
        indice += 1
        importe = None
        while indice < len(lineas):
            if re.match(r'(?i)^subtotal', lineas[indice]):
                break
            if cantidad.fullmatch(lineas[indice]) and indice + 1 < len(lineas) and re.search(r'(?i)(pieza|servicio|unidad|H87|ACT|E48)', lineas[indice + 1]):
                break
            coincidencia = numero.fullmatch(lineas[indice])
            if coincidencia:
                importe = str(_decimal_factura(coincidencia.group(1)))
            indice += 1
        texto_descripcion = ' '.join(descripcion).strip()
        if texto_descripcion:
            conceptos.append({'cantidad': cantidad_texto, 'descripcion': texto_descripcion, 'precio_unitario': precio, 'importe': importe or str(_decimal_factura(cantidad_texto) * _decimal_factura(precio))})
    return conceptos


def _extraer_factura_pdf(ruta):
    try:
        import fitz
        documento = fitz.open(ruta)
        texto = '\n'.join(pagina.get_text('text') for pagina in documento)
        documento.close()
    except Exception as exc:
        raise ValueError('No fue posible leer el PDF de la factura.') from exc
    texto = texto.strip()
    for incorrecto, correcto in {
        'Delegaci�n': 'Delegación', 'M�xico': 'México', 'Le�n': 'León',
        'emisi�n': 'emisión', 'Descripci�n': 'Descripción', 'R�gimen': 'Régimen',
        'cr�dito': 'crédito', 'N�mero': 'Número', 'p�gina': 'página',
    }.items():
        texto = texto.replace(incorrecto, correcto)
    if len(texto) < 40:
        raise ValueError('El PDF parece escaneado y no contiene texto reconocible. Adjunta también el XML CFDI.')
    rfc_encontrados = [rfc for rfc in re.findall(r'\b[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}\b', texto.upper()) if rfc != 'PBL890425149']
    fecha = datetime.now().date().isoformat()
    for patron, formato in [(r'\b(\d{2}/\d{2}/20\d{2})\b', '%d/%m/%Y'), (r'\b(20\d{2}-\d{2}-\d{2})\b', '%Y-%m-%d')]:
        for candidato in re.findall(patron, texto):
            try:
                fecha = datetime.strptime(candidato, formato).date().isoformat()
                break
            except ValueError:
                continue
        if fecha != datetime.now().date().isoformat():
            break
    proveedor = ''
    for patron in [r'(?im)^\s*(?:EMISOR|NOMBRE|RAZ[ÓO]N SOCIAL)\s*:?\s*(.+)$', r'(?im)^\s*PROVEEDOR\s*:?\s*(.+)$']:
        encontrado = re.search(patron, texto)
        if encontrado:
            proveedor = encontrado.group(1).strip(' :-').upper()[:160]
            break
    lineas = [l.strip() for l in texto.splitlines() if l.strip()]
    telefono = ''
    domicilio = ''
    for indice, linea in enumerate(lineas):
        telefono_match = re.search(r'(?i)\bTel(?:éfono)?\.?\s*:?[ ]*([0-9 ()+\-]{7,}?)(?=\s*Fax|$)', linea)
        if not telefono_match:
            continue
        telefono = ' '.join(telefono_match.group(1).split()).strip()
        direccion = []
        candidato_proveedor = ''
        for anterior in reversed(lineas[max(0, indice - 6):indice]):
            letras = ''.join(c for c in anterior if c.isalpha())
            if direccion and letras and anterior == anterior.upper():
                candidato_proveedor = anterior
                break
            direccion.insert(0, anterior)
        domicilio = ', '.join(direccion[-4:])[:300]
        if not proveedor and candidato_proveedor:
            proveedor = candidato_proveedor[:160]
        break
    conceptos_pdf = _extraer_conceptos_pdf(lineas)
    lineas_utiles = [l for l in lineas if 8 <= len(l) <= 220]
    descripcion = ' | '.join(lineas_utiles[:8])[:1800]
    subtotal = _buscar_importe_pdf(texto, 'SUBTOTAL')
    subtotal_separado = re.search(r'(?is)Subtotal.{0,220}?\$\s*([\d,]+\.\d{2})', texto)
    if _decimal_factura(subtotal) <= 0 and subtotal_separado:
        subtotal = str(_decimal_factura(subtotal_separado.group(1)))
    iva = _buscar_importe_pdf(texto, r'(?:IVA|I\.V\.A\.)')
    iva_traslado = re.search(r'(?is)IVA\s+Traslado\s+Tasa\s*\([^)]+\)\s*:\s*([\d,]+\.\d{2})', texto)
    if iva_traslado:
        iva = str(_decimal_factura(iva_traslado.group(1)))
    for linea in lineas if 'lineas' in locals() else texto.splitlines():
        if 'IVA' in linea.upper():
            importes_iva = re.findall(r'([\d,]+\.\d{2})', linea)
            if len(importes_iva) >= 2:
                iva = str(_decimal_factura(importes_iva[-1]))
                break
    total = _buscar_importe_pdf(texto, 'TOTAL')
    if _decimal_factura(subtotal) <= 0 and _decimal_factura(total) > 0 and _decimal_factura(iva) > 0:
        subtotal = str(_decimal_factura(total) - _decimal_factura(iva))
    return {
        'origen': 'PDF', 'proveedor': proveedor, 'rfc': rfc_encontrados[0] if rfc_encontrados else '',
        'fecha': fecha, 'subtotal': subtotal, 'iva': iva, 'total': total, 'moneda': 'MXN',
        'domicilio': domicilio, 'telefono': telefono,
        'conceptos': conceptos_pdf or [{'cantidad': '1', 'descripcion': descripcion, 'precio_unitario': subtotal, 'importe': subtotal}],
    }


def _sugerir_partida_factura(descripcion, partidas):
    palabras = set(re.findall(r'[a-záéíóúñ]{4,}', (descripcion or '').lower()))
    ignorar = {'para', 'como', 'este', 'esta', 'servicio', 'material', 'gastos', 'pago'}
    palabras -= ignorar
    mejor, puntuacion_mejor = None, 0
    for partida in partidas:
        catalogo = f'{partida.nombre} {partida.descripcion or ""}'.lower()
        puntuacion = sum(1 for palabra in palabras if palabra in catalogo)
        if puntuacion > puntuacion_mejor:
            mejor, puntuacion_mejor = partida, puntuacion
    return mejor


def _normalizar_nombre_proveedor(nombre):
    nombre = (nombre or '').upper().replace('&', ' Y ')
    nombre = re.sub(r'[^A-Z0-9Ñ ]+', ' ', nombre)
    nombre = re.sub(r'\b(SA|S A|DE|CV|C V|SAPI|S DE RL|SOCIEDAD ANONIMA)\b', ' ', nombre)
    return ' '.join(nombre.split())


def _reconocer_proveedor_factura(nombre, rfc):
    if rfc:
        encontrado = ProveedorCompra.query.filter(
            db.func.upper(ProveedorCompra.rfc) == rfc.upper(),
            ProveedorCompra.activo.is_(True),
        ).first()
        if encontrado:
            return encontrado
    buscado = _normalizar_nombre_proveedor(nombre)
    if not buscado:
        return None
    for proveedor in ProveedorCompra.query.filter_by(activo=True).all():
        guardado = _normalizar_nombre_proveedor(proveedor.nombre)
        if guardado == buscado or (len(buscado) >= 8 and (buscado in guardado or guardado in buscado)):
            return proveedor
    return None


def _guardar_factura_pendiente(archivo, token):
    nombre = secure_filename(archivo.filename)
    extension = os.path.splitext(nombre)[1].lower()
    if extension not in {'.pdf', '.xml'}:
        raise ValueError('Para generar la orden sólo se permiten archivos PDF o XML.')
    archivo.stream.seek(0, os.SEEK_END)
    tamano = archivo.stream.tell()
    archivo.stream.seek(0)
    if tamano > 15 * 1024 * 1024:
        raise ValueError(f'El archivo {nombre} excede el límite de 15 MB.')
    carpeta = os.path.join(current_app.config['UPLOAD_FOLDER'], 'facturas_pendientes')
    os.makedirs(carpeta, exist_ok=True)
    nombre_interno = f'{token}_{secrets.token_hex(6)}{extension}'
    ruta = os.path.join(carpeta, nombre_interno)
    archivo.save(ruta)
    return {'original': nombre, 'interno': nombre_interno, 'extension': extension, 'mime': archivo.mimetype, 'tamano': tamano, 'ruta': ruta}


@main.route('/ordenes-compra')
@login_required
@acceso_compras_requerido
def ordenes_compra():
    _areas_compra_iniciales()
    areas = _areas_compra_usuario()
    ids_area = [area.id for area in areas]
    area_id = request.args.get('area', type=int)
    consulta = OrdenCompra.query.filter(OrdenCompra.area_id.in_(ids_area))
    if area_id in ids_area:
        consulta = consulta.filter_by(area_id=area_id)
    ordenes = consulta.order_by(OrdenCompra.creado_en.desc()).limit(100).all()
    return render_template(
        'ordenes_compra.html',
        ordenes=ordenes,
        areas=areas,
        area_seleccionada=area_id,
    )


@main.route('/ordenes-compra/catalogos')
@login_required
@acceso_compras_requerido
def catalogos_compra():
    return render_template(
        'catalogos_compra.html',
        proveedores=ProveedorCompra.query.order_by(ProveedorCompra.nombre).all(),
        partidas=PartidaPresupuestal.query.order_by(PartidaPresupuestal.codigo).all(),
    )


@main.route('/ordenes-compra/proveedores', methods=['POST'])
@login_required
@acceso_compras_requerido
def guardar_proveedor_compra():
    nombre = request.form.get('nombre', '').strip().upper()
    if not nombre:
        flash('El nombre del proveedor es obligatorio.', 'warning')
        return redirect(url_for('main.catalogos_compra'))
    proveedor = ProveedorCompra.query.filter(db.func.lower(ProveedorCompra.nombre) == nombre.lower()).first()
    if not proveedor:
        proveedor = ProveedorCompra(nombre=nombre)
        db.session.add(proveedor)
    proveedor.domicilio = request.form.get('domicilio', '').strip() or None
    proveedor.atencion_a = request.form.get('atencion_a', '').strip() or None
    proveedor.telefono = request.form.get('telefono', '').strip() or None
    proveedor.rfc = request.form.get('rfc', '').strip().upper() or None
    proveedor.correo = request.form.get('correo', '').strip().lower() or None
    proveedor.activo = True
    db.session.commit()
    flash(f'Proveedor {nombre} guardado.', 'success')
    return redirect(url_for('main.catalogos_compra'))


@main.route('/ordenes-compra/partidas-presupuestales', methods=['POST'])
@login_required
@acceso_compras_requerido
def guardar_partida_presupuestal():
    codigo = request.form.get('codigo', '').strip().upper()
    nombre = request.form.get('nombre', '').strip().upper()
    if not codigo or not nombre:
        flash('Código y nombre de la partida son obligatorios.', 'warning')
        return redirect(url_for('main.catalogos_compra'))
    partida = PartidaPresupuestal.query.filter(db.func.lower(PartidaPresupuestal.codigo) == codigo.lower()).first()
    if not partida:
        partida = PartidaPresupuestal(codigo=codigo, nombre=nombre)
        db.session.add(partida)
    partida.nombre = nombre
    partida.descripcion = request.form.get('descripcion', '').strip() or None
    partida.activa = True
    db.session.commit()
    flash(f'Partida {codigo} guardada.', 'success')
    return redirect(url_for('main.catalogos_compra'))


@main.route('/ordenes-compra/areas', methods=['POST'])
@login_required
@acceso_compras_requerido
def crear_area_compra():
    if current_user.username.lower() != 'admin':
        return jsonify({'error': 'Solo administración puede configurar áreas.'}), 403
    nombre = request.form.get('nombre', '').strip().upper()
    codigo = request.form.get('codigo', '').strip().upper()
    try:
        consecutivo_inicial = int(request.form.get('consecutivo_inicial', '0'))
        if not nombre or not codigo.isalnum() or len(codigo) > 12 or consecutivo_inicial < 0:
            raise ValueError('Indica nombre, código alfanumérico y último consecutivo válido.')
        db.session.add(AreaCompra(nombre=nombre, codigo=codigo, ultimo_consecutivo=consecutivo_inicial))
        db.session.commit()
        flash(f'Área {nombre} configurada. El siguiente folio será PBL/{codigo}/{consecutivo_inicial + 1:03d}.', 'success')
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), 'warning')
    except Exception:
        db.session.rollback()
        flash('El nombre o código de esa área ya existe.', 'warning')
    return redirect(url_for('main.ordenes_compra'))


@main.route('/ordenes-compra/desde-factura', methods=['GET', 'POST'])
@login_required
@acceso_compras_requerido
def orden_desde_factura():
    _areas_compra_iniciales()
    areas = _areas_compra_usuario()
    partidas = PartidaPresupuestal.query.filter_by(activa=True).order_by(PartidaPresupuestal.codigo).all()
    if request.method == 'GET':
        return render_template('orden_desde_factura.html', etapa='carga', areas=areas, partidas=partidas)

    accion = request.form.get('accion', 'analizar')
    carpeta_pendiente = os.path.join(current_app.config['UPLOAD_FOLDER'], 'facturas_pendientes')
    os.makedirs(carpeta_pendiente, exist_ok=True)
    if accion == 'analizar':
        guardados = []
        try:
            archivos = [a for a in (request.files.get('pdf'), request.files.get('xml')) if a and a.filename]
            if not archivos:
                raise ValueError('Adjunta el PDF de la factura o su XML CFDI.')
            token = secrets.token_urlsafe(18)
            guardados = [_guardar_factura_pendiente(archivo, token) for archivo in archivos]
            xml = next((a for a in guardados if a['extension'] == '.xml'), None)
            pdf = next((a for a in guardados if a['extension'] == '.pdf'), None)
            datos_pdf = _extraer_factura_pdf(pdf['ruta']) if pdf else {}
            datos = _extraer_factura_xml(xml['ruta']) if xml else datos_pdf
            if xml and datos_pdf:
                datos['domicilio'] = datos_pdf.get('domicilio', '')
                datos['telefono'] = datos_pdf.get('telefono', '')
            if not datos['conceptos']:
                raise ValueError('No se encontraron conceptos en la factura.')
            descripcion = '\n'.join(item['descripcion'] for item in datos['conceptos'])[:3000]
            subtotal = _decimal_factura(datos.get('subtotal'))
            if subtotal <= 0:
                subtotal = sum((_decimal_factura(item.get('importe')) for item in datos['conceptos']), Decimal('0'))
            proveedor = _reconocer_proveedor_factura(datos.get('proveedor'), datos.get('rfc'))
            sugerida = _sugerir_partida_factura(descripcion, partidas)
            meta = {
                'token': token,
                'archivos': [{k: v for k, v in archivo.items() if k != 'ruta'} for archivo in guardados],
                'datos': datos,
            }
            with open(os.path.join(carpeta_pendiente, f'{token}.json'), 'w', encoding='utf-8') as salida:
                pyjson.dump(meta, salida, ensure_ascii=False)
            revision = {
                'token': token, 'origen': datos['origen'], 'proveedor_id': proveedor.id if proveedor else '',
                'proveedor': (proveedor.nombre if proveedor else datos.get('proveedor')) or 'NO INDICA',
                'rfc': (proveedor.rfc if proveedor and proveedor.rfc else datos.get('rfc')) or 'NO INDICA',
                'domicilio': (proveedor.domicilio if proveedor else None) or datos.get('domicilio') or 'NO INDICA',
                'atencion_a': (proveedor.atencion_a if proveedor else None) or 'NO INDICA',
                'telefono': (proveedor.telefono if proveedor else None) or datos.get('telefono') or 'NO INDICA',
                'fecha': datetime.now().date().isoformat(),
                'descripcion': descripcion, 'conceptos': datos['conceptos'],
                'subtotal': str(subtotal),
                'total_factura': datos.get('total', '0'), 'iva_factura': datos.get('iva', '0'),
                'partida_codigo': sugerida.codigo if sugerida else '',
                'justificacion': f"ADQUISICIÓN DE {datos['conceptos'][0]['descripcion']} SEGÚN FACTURA"[:1000].upper(),
            }
            return render_template('orden_desde_factura.html', etapa='revision', areas=areas, partidas=partidas, revision=revision)
        except (ValueError, ET.ParseError) as exc:
            for archivo in guardados:
                try:
                    os.remove(archivo['ruta'])
                except OSError:
                    pass
            flash(str(exc), 'warning')
        except Exception:
            current_app.logger.exception('Error analizando factura')
            flash('No fue posible analizar la factura. Si tienes el XML CFDI, adjúntalo junto con el PDF.', 'danger')
        return render_template('orden_desde_factura.html', etapa='carga', areas=areas, partidas=partidas), 400

    token = request.form.get('token', '')
    if not re.fullmatch(r'[A-Za-z0-9_-]{20,40}', token):
        flash('La revisión de la factura ya no es válida. Vuelve a cargarla.', 'warning')
        return redirect(url_for('main.orden_desde_factura'))
    ruta_meta = os.path.join(carpeta_pendiente, f'{token}.json')
    try:
        with open(ruta_meta, encoding='utf-8') as entrada:
            meta = pyjson.load(entrada)
        if meta.get('token') != token:
            raise ValueError('La factura pendiente no coincide con esta revisión.')
        area_id = request.form.get('area_id', type=int)
        if area_id not in {area.id for area in areas}:
            raise ValueError('El área seleccionada no está autorizada para tu usuario.')
        partida = PartidaPresupuestal.query.filter_by(codigo=request.form.get('partida_codigo', '').strip(), activa=True).first()
        if not partida:
            raise ValueError('Selecciona una partida presupuestal válida.')
        nombre_proveedor = request.form.get('proveedor', '').strip().upper() or 'NO INDICA'
        rfc = request.form.get('rfc', '').strip().upper() or 'NO INDICA'
        proveedor = db.session.get(ProveedorCompra, request.form.get('proveedor_id', type=int))
        if not proveedor and rfc != 'NO INDICA':
            proveedor = ProveedorCompra.query.filter(db.func.upper(ProveedorCompra.rfc) == rfc).first()
        if not proveedor:
            proveedor = ProveedorCompra.query.filter(db.func.lower(ProveedorCompra.nombre) == nombre_proveedor.lower()).first()
        if not proveedor:
            proveedor = ProveedorCompra(nombre=nombre_proveedor)
            db.session.add(proveedor)
        proveedor.nombre = nombre_proveedor
        proveedor.rfc = rfc if rfc != 'NO INDICA' else (proveedor.rfc or 'NO INDICA')
        proveedor.domicilio = request.form.get('domicilio', '').strip() or proveedor.domicilio or 'NO INDICA'
        proveedor.atencion_a = request.form.get('atencion_a', '').strip() or proveedor.atencion_a or 'NO INDICA'
        proveedor.telefono = request.form.get('telefono', '').strip() or proveedor.telefono or 'NO INDICA'
        proveedor.activo = True
        fecha = datetime.now().date()
        fecha_entrega = datetime.strptime(request.form.get('fecha_entrega_requerida', ''), '%Y-%m-%d').date()
        cantidades = request.form.getlist('cantidad[]')
        descripciones = request.form.getlist('descripcion[]')
        precios = request.form.getlist('precio_unitario[]')
        conceptos = []
        for posicion, (cantidad_texto, descripcion, precio_texto) in enumerate(zip(cantidades, descripciones, precios), 1):
            cantidad = _decimal_factura(cantidad_texto)
            precio_unitario = _decimal_factura(precio_texto)
            descripcion = descripcion.strip()
            if descripcion and cantidad > 0 and precio_unitario > 0:
                conceptos.append((posicion, cantidad, descripcion, precio_unitario))
        justificacion = request.form.get('justificacion', '').strip()
        if not conceptos or not justificacion:
            raise ValueError('Confirma la cantidad, descripción y precio unitario de cada artículo.')
        siguiente = db.session.execute(db.text(
            'UPDATE area_compra SET ultimo_consecutivo = ultimo_consecutivo + 1 '
            'WHERE id = :area_id AND activa = 1 RETURNING ultimo_consecutivo, codigo'
        ), {'area_id': area_id}).first()
        if not siguiente:
            raise ValueError('Selecciona un área válida.')
        consecutivo, codigo_area = siguiente
        orden = OrdenCompra(
            area_id=area_id, consecutivo=consecutivo, folio=f'PBL/{codigo_area}/{consecutivo:03d}',
            solicitante_id=current_user.id, fecha=fecha, fecha_entrega_requerida=fecha_entrega,
            proveedor_catalogo=proveedor, partida_presupuestal=partida, proveedor=proveedor.nombre,
            domicilio=proveedor.domicilio, atencion_a=proveedor.atencion_a, telefono=proveedor.telefono,
            cuenta_presupuestal=f'{partida.codigo} {partida.nombre}', proyecto_programa=partida.nombre.upper(),
            fuente_financiamiento='RECURSOS PROPIOS', tipo_compra='ADQUISICIÓN DIRECTA',
            justificacion=justificacion, iva_porcentaje=Decimal('16'),
            estado='DESDE FACTURA',
        )
        db.session.add(orden)
        db.session.flush()
        for posicion, cantidad, descripcion, precio_unitario in conceptos:
            db.session.add(PartidaOrdenCompra(orden_id=orden.id, posicion=posicion, cantidad=cantidad, descripcion=descripcion, precio_unitario=precio_unitario))
        carpeta_final = os.path.join(current_app.config['UPLOAD_FOLDER'], 'facturas')
        os.makedirs(carpeta_final, exist_ok=True)
        for archivo in meta.get('archivos', []):
            if not archivo.get('interno', '').startswith(f'{token}_'):
                continue
            origen = os.path.join(carpeta_pendiente, archivo['interno'])
            if not os.path.isfile(origen):
                continue
            nombre_final = f'{orden.id}_{secrets.token_hex(12)}{archivo["extension"]}'
            shutil.move(origen, os.path.join(carpeta_final, nombre_final))
            db.session.add(FacturaOrdenCompra(
                orden_id=orden.id, nombre_original=archivo['original'], nombre_archivo=nombre_final,
                tipo_mime=archivo.get('mime'), tamano=archivo.get('tamano', 0), subido_por_id=current_user.id,
            ))
        db.session.commit()
        try:
            os.remove(ruta_meta)
        except OSError:
            pass
        flash(f'Orden {orden.folio} generada desde la factura.', 'success')
        return redirect(url_for('main.ver_orden_compra', orden_id=orden.id))
    except (ValueError, InvalidOperation) as exc:
        db.session.rollback()
        flash(str(exc), 'warning')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error generando orden desde factura')
        flash('No fue posible generar la orden desde la factura.', 'danger')
    return redirect(url_for('main.orden_desde_factura'))


@main.route('/ordenes-compra/nueva', methods=['GET', 'POST'])
@login_required
@acceso_compras_requerido
def nueva_orden_compra():
    _areas_compra_iniciales()
    areas = _areas_compra_usuario()
    proveedores = ProveedorCompra.query.filter_by(activo=True).order_by(ProveedorCompra.nombre).all()
    partidas_catalogo = PartidaPresupuestal.query.filter_by(activa=True).order_by(PartidaPresupuestal.codigo).all()
    if request.method == 'GET':
        return render_template('orden_compra_form.html', areas=areas, proveedores=proveedores, partidas_catalogo=partidas_catalogo, hoy=datetime.now().date())

    try:
        area_id = int(request.form.get('area_id', ''))
        if area_id not in {area.id for area in areas}:
            raise ValueError('El área seleccionada no está autorizada para tu usuario.')
        fecha = datetime.strptime(request.form.get('fecha', ''), '%Y-%m-%d').date()
        fecha_entrega = datetime.strptime(request.form.get('fecha_entrega_requerida', ''), '%Y-%m-%d').date()
        proveedor_catalogo = db.session.get(ProveedorCompra, request.form.get('proveedor_id', type=int))
        codigo_partida = request.form.get('partida_codigo', '').strip()
        partida_catalogo = PartidaPresupuestal.query.filter(
            db.func.lower(PartidaPresupuestal.codigo) == codigo_partida.lower(),
            PartidaPresupuestal.activa.is_(True),
        ).first()
        if not proveedor_catalogo or not proveedor_catalogo.activo:
            raise ValueError('Selecciona un proveedor guardado en el catálogo.')
        if not partida_catalogo or not partida_catalogo.activa:
            raise ValueError('Selecciona una partida presupuestal guardada en el catálogo.')
        cantidades = request.form.getlist('cantidad[]')[:1]
        descripciones = request.form.getlist('descripcion[]')[:1]
        precios = request.form.getlist('precio_unitario[]')[:1]
        partidas = []
        for posicion, (cantidad, descripcion, precio) in enumerate(zip(cantidades, descripciones, precios), 1):
            descripcion = descripcion.strip()
            if not descripcion and not cantidad and not precio:
                continue
            cantidad_decimal = Decimal(cantidad)
            precio_decimal = Decimal(precio)
            if cantidad_decimal <= 0 or precio_decimal < 0 or not descripcion:
                raise ValueError('Cada partida requiere descripción, cantidad mayor a cero y precio válido.')
            partidas.append((posicion, cantidad_decimal, descripcion, precio_decimal))
        if not partidas:
            raise ValueError('Agrega al menos una partida a la orden.')

        # RETURNING reserva un número único incluso con capturas simultáneas.
        siguiente = db.session.execute(
            db.text(
                'UPDATE area_compra SET ultimo_consecutivo = ultimo_consecutivo + 1 '
                'WHERE id = :area_id AND activa = 1 RETURNING ultimo_consecutivo, codigo'
            ),
            {'area_id': area_id},
        ).first()
        if not siguiente:
            raise ValueError('El área seleccionada no existe o está inactiva.')
        consecutivo, codigo_area = siguiente
        folio = f'PBL/{codigo_area}/{consecutivo:03d}'
        orden = OrdenCompra(
            area_id=area_id,
            consecutivo=consecutivo,
            folio=folio,
            solicitante_id=current_user.id,
            fecha=fecha,
            fecha_entrega_requerida=fecha_entrega,
            proveedor_id=proveedor_catalogo.id,
            partida_presupuestal_id=partida_catalogo.id,
            proveedor=proveedor_catalogo.nombre,
            domicilio=proveedor_catalogo.domicilio or 'NO INDICA',
            atencion_a=proveedor_catalogo.atencion_a or 'NO INDICA',
            telefono=proveedor_catalogo.telefono or 'NO INDICA',
            cuenta_presupuestal=f'{partida_catalogo.codigo} {partida_catalogo.nombre}',
            proyecto_programa=partida_catalogo.nombre.upper(),
            fuente_financiamiento='RECURSOS PROPIOS',
            tipo_compra='ADQUISICIÓN DIRECTA',
            justificacion=request.form.get('justificacion', '').strip(),
            iva_porcentaje=Decimal('16'),
        )
        if not orden.proveedor or not orden.cuenta_presupuestal or not orden.justificacion:
            raise ValueError('Proveedor, cuenta presupuestal y justificación son obligatorios.')
        db.session.add(orden)
        db.session.flush()
        for posicion, cantidad, descripcion, precio in partidas:
            db.session.add(PartidaOrdenCompra(
                orden_id=orden.id,
                posicion=posicion,
                cantidad=cantidad,
                descripcion=descripcion,
                precio_unitario=precio,
            ))
        db.session.commit()
        flash(f'Orden {folio} creada correctamente.', 'success')
        return redirect(url_for('main.ver_orden_compra', orden_id=orden.id))
    except (ValueError, InvalidOperation) as exc:
        db.session.rollback()
        flash(str(exc), 'warning')
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error creando orden de compra')
        flash('No fue posible guardar la orden. Intenta nuevamente.', 'danger')
    return render_template('orden_compra_form.html', areas=areas, proveedores=proveedores, partidas_catalogo=partidas_catalogo, hoy=datetime.now().date()), 400


@main.route('/ordenes-compra/<int:orden_id>')
@login_required
@acceso_compras_requerido
def ver_orden_compra(orden_id):
    orden = db.get_or_404(OrdenCompra, orden_id)
    if orden.area_id not in {area.id for area in _areas_compra_usuario()}:
        flash('No tienes permiso para consultar órdenes de esa área.', 'warning')
        return redirect(url_for('main.ordenes_compra'))
    factura_paginas = []
    if orden.estado == 'DESDE FACTURA':
        try:
            import fitz
            carpeta = os.path.join(current_app.config['UPLOAD_FOLDER'], 'facturas')
            for factura in orden.facturas:
                if os.path.splitext(factura.nombre_archivo)[1].lower() != '.pdf':
                    continue
                ruta = os.path.join(carpeta, factura.nombre_archivo)
                with fitz.open(ruta) as documento:
                    factura_paginas.extend((factura, pagina) for pagina in range(documento.page_count))
        except Exception:
            current_app.logger.exception('No fue posible preparar la factura para impresión')
    return render_template('orden_compra_documento.html', orden=orden, factura_paginas=factura_paginas)


@main.route('/ordenes-compra/facturas/<int:factura_id>/pagina/<int:pagina>')
@login_required
@acceso_compras_requerido
def pagina_factura_compra(factura_id, pagina):
    factura = db.get_or_404(FacturaOrdenCompra, factura_id)
    if factura.orden.area_id not in {area.id for area in _areas_compra_usuario()}:
        return 'No autorizado.', 403
    if os.path.splitext(factura.nombre_archivo)[1].lower() != '.pdf':
        return 'La factura no es un PDF.', 404
    ruta = os.path.join(current_app.config['UPLOAD_FOLDER'], 'facturas', factura.nombre_archivo)
    try:
        import fitz
        with fitz.open(ruta) as documento:
            if pagina < 0 or pagina >= documento.page_count:
                return 'Página no encontrada.', 404
            pixmap = documento[pagina].get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            imagen = BytesIO(pixmap.tobytes('png'))
        respuesta = send_file(imagen, mimetype='image/png')
        respuesta.headers['Cache-Control'] = 'private, max-age=3600'
        return respuesta
    except Exception:
        current_app.logger.exception('No fue posible renderizar una página de factura')
        return 'No fue posible mostrar la factura.', 500


@main.route('/ordenes-compra/<int:orden_id>/facturas', methods=['POST'])
@login_required
@acceso_compras_requerido
def adjuntar_factura_compra(orden_id):
    orden = db.get_or_404(OrdenCompra, orden_id)
    if orden.area_id not in {area.id for area in _areas_compra_usuario()}:
        flash('No tienes permiso para modificar órdenes de esa área.', 'warning')
        return redirect(url_for('main.ordenes_compra'))
    archivo = request.files.get('factura')
    if not archivo or not archivo.filename:
        flash('Selecciona una factura para adjuntar.', 'warning')
        return redirect(url_for('main.ver_orden_compra', orden_id=orden.id))
    extension = os.path.splitext(secure_filename(archivo.filename))[1].lower()
    if extension not in {'.pdf', '.xml', '.png', '.jpg', '.jpeg'}:
        flash('Formato no permitido. Usa PDF, XML, PNG o JPG.', 'warning')
        return redirect(url_for('main.ver_orden_compra', orden_id=orden.id))
    archivo.stream.seek(0, os.SEEK_END)
    tamano = archivo.stream.tell()
    archivo.stream.seek(0)
    if tamano > 15 * 1024 * 1024:
        flash('La factura excede el límite de 15 MB.', 'warning')
        return redirect(url_for('main.ver_orden_compra', orden_id=orden.id))
    nombre_archivo = f'{orden.id}_{secrets.token_hex(12)}{extension}'
    carpeta = os.path.join(current_app.config['UPLOAD_FOLDER'], 'facturas')
    os.makedirs(carpeta, exist_ok=True)
    archivo.save(os.path.join(carpeta, nombre_archivo))
    db.session.add(FacturaOrdenCompra(
        orden_id=orden.id,
        nombre_original=secure_filename(archivo.filename),
        nombre_archivo=nombre_archivo,
        tipo_mime=archivo.mimetype,
        tamano=tamano,
        subido_por_id=current_user.id,
    ))
    db.session.commit()
    flash('Factura adjuntada correctamente.', 'success')
    return redirect(url_for('main.ver_orden_compra', orden_id=orden.id))


@main.route('/ordenes-compra/facturas/<int:factura_id>')
@login_required
@acceso_compras_requerido
def descargar_factura_compra(factura_id):
    factura = db.get_or_404(FacturaOrdenCompra, factura_id)
    if factura.orden.area_id not in {area.id for area in _areas_compra_usuario()}:
        return 'No autorizado.', 403
    carpeta = os.path.join(current_app.config['UPLOAD_FOLDER'], 'facturas')
    return send_from_directory(carpeta, factura.nombre_archivo, as_attachment=True, download_name=factura.nombre_original)
@main.route('/uploads/<filename>')
@login_required
def serve_uploaded_file(filename):
    upload_folder = current_app.config['UPLOAD_FOLDER']
    return send_from_directory(upload_folder, filename)


@main.route('/sw.js')
def service_worker():
    response = send_from_directory(os.path.join(current_app.root_path, 'static', 'js'), 'sw.js')
    response.headers['Service-Worker-Allowed'] = '/'
    response.headers['Cache-Control'] = 'no-cache'
    return response

@main.route('/chat-widget')
@login_required
def chat_widget():
    return render_template('chat_widget.html')

@main.route('/login', methods=['GET', 'POST'])
def login():

    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    form = LoginForm()
    if current_app.config.get('LOCAL_DEVELOPMENT'):
        # RecaptchaField valida directamente contra Google; en el modo local
        # explícito retiramos únicamente ese validador. CSRF y usuario siguen
        # validándose normalmente.
        form.recaptcha.validators = []
    if form.validate_on_submit():
        username = form.username.data.strip()
        user = User.query.filter(db.func.lower(User.username) == username.lower()).first()
        if not user and current_app.config.get('LOCAL_DEVELOPMENT'):
            user = User(username=username, nombre=f'Usuario local ({username})')
            user.set_password(secrets.token_urlsafe(24))
            db.session.add(user)
            db.session.commit()
            current_app.logger.info('Usuario temporal creado para pruebas locales: %s', username)
        if user:
            if user.username == 'admin':
                if not form.password.data or not user.check_password(form.password.data):
                    flash('Contraseña incorrecta.', 'danger')
                    return render_template('login.html', form=form)
            login_user(user)
            flash('Inicio de sesión exitoso.', 'success')
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            return redirect(url_for('main.dashboard'))
        else:
            flash('Usuario incorrecto.', 'danger')
    elif request.method == 'POST':
        current_app.logger.warning('Inicio de sesión rechazado por formulario: %s', form.errors)
        flash('Revisa el usuario y vuelve a intentarlo.', 'warning')

    return render_template('login.html', form=form)


def _passkey_context():
    host = request.host.split(':', 1)[0].lower()
    origin = f"http://{request.host}" if host in {'localhost', '127.0.0.1'} else f"https://{host}"
    return host, origin


def _b64url_encode(value):
    return base64.urlsafe_b64encode(value).rstrip(b'=').decode('ascii')


def _b64url_decode(value):
    return base64.urlsafe_b64decode(value + '=' * (-len(value) % 4))


@main.route('/configurar-passkey')
@login_required
def configurar_passkey():
    if current_user.username.lower() in USUARIOS_SIN_FUNCIONES:
        return redirect(url_for('main.dashboard'))
    credenciales = PasskeyCredential.query.filter_by(user_id=current_user.id).order_by(
        PasskeyCredential.creado_en.desc()
    ).all()
    return render_template('configurar_passkey.html', credenciales=credenciales)


@main.route('/api/passkey/register/options', methods=['POST'])
@login_required
def passkey_register_options():
    if current_user.username.lower() in USUARIOS_SIN_FUNCIONES:
        return jsonify({'success': False, 'error': 'Usuario no autorizado'}), 403
    rp_id, _ = _passkey_context()
    existentes = PasskeyCredential.query.filter_by(user_id=current_user.id).all()
    options = generate_registration_options(
        rp_id=rp_id,
        rp_name='Intranet Bomberos de Leon',
        user_id=f'intranet-user-{current_user.id}'.encode(),
        user_name=current_user.username,
        user_display_name=current_user.nombre,
        exclude_credentials=[PublicKeyCredentialDescriptor(id=item.credential_id) for item in existentes],
        authenticator_selection=AuthenticatorSelectionCriteria(
            resident_key=ResidentKeyRequirement.REQUIRED,
            require_resident_key=True,
            user_verification=UserVerificationRequirement.REQUIRED,
        ),
        timeout=60000,
    )
    session['passkey_registration_challenge'] = _b64url_encode(options.challenge)
    return current_app.response_class(options_to_json(options), mimetype='application/json')


@main.route('/api/passkey/register/verify', methods=['POST'])
@login_required
def passkey_register_verify():
    challenge = session.pop('passkey_registration_challenge', None)
    if not challenge:
        return jsonify({'success': False, 'error': 'El registro expiro. Intenta nuevamente'}), 403
    rp_id, origin = _passkey_context()
    try:
        verification = verify_registration_response(
            credential=request.get_json(force=True),
            expected_challenge=_b64url_decode(challenge),
            expected_rp_id=rp_id,
            expected_origin=origin,
            require_user_verification=True,
        )
    except Exception as exc:
        logging.getLogger(__name__).warning('Registro passkey rechazado: %s', exc)
        return jsonify({'success': False, 'error': 'No se pudo validar la passkey'}), 400

    if PasskeyCredential.query.filter_by(credential_id=verification.credential_id).first():
        return jsonify({'success': False, 'error': 'Este dispositivo ya esta registrado'}), 409
    body = request.get_json(silent=True) or {}
    transports = ((body.get('response') or {}).get('transports') or [])
    credential = PasskeyCredential(
        user_id=current_user.id,
        credential_id=verification.credential_id,
        public_key=verification.credential_public_key,
        sign_count=verification.sign_count,
        nombre_dispositivo=(body.get('deviceName') or 'Celular personal')[:100],
        transports=','.join(transports)[:100],
    )
    db.session.add(credential)
    db.session.commit()
    return jsonify({'success': True})


@main.route('/api/passkey/auth/options', methods=['POST'])
def passkey_auth_options():
    rp_id, _ = _passkey_context()
    options = generate_authentication_options(
        rp_id=rp_id,
        allow_credentials=[],
        user_verification=UserVerificationRequirement.REQUIRED,
        timeout=60000,
    )
    session['passkey_auth_challenge'] = _b64url_encode(options.challenge)
    return current_app.response_class(options_to_json(options), mimetype='application/json')


@main.route('/api/passkey/auth/verify', methods=['POST'])
def passkey_auth_verify():
    challenge = session.pop('passkey_auth_challenge', None)
    body = request.get_json(silent=True) or {}
    if not challenge:
        return jsonify({'success': False, 'error': 'La solicitud expiro. Intenta nuevamente'}), 403
    try:
        credential_id = _b64url_decode(body.get('id', ''))
    except Exception:
        return jsonify({'success': False, 'error': 'Credencial invalida'}), 400
    stored = PasskeyCredential.query.filter_by(credential_id=credential_id).first()
    if not stored:
        return jsonify({'success': False, 'error': 'Este dispositivo no esta vinculado'}), 404
    rp_id, origin = _passkey_context()
    try:
        verification = verify_authentication_response(
            credential=body,
            expected_challenge=_b64url_decode(challenge),
            expected_rp_id=rp_id,
            expected_origin=origin,
            credential_public_key=stored.public_key,
            credential_current_sign_count=stored.sign_count,
            require_user_verification=True,
        )
    except Exception as exc:
        logging.getLogger(__name__).warning('Acceso passkey rechazado: %s', exc)
        return jsonify({'success': False, 'error': 'No se pudo comprobar tu identidad'}), 403

    stored.sign_count = verification.new_sign_count
    stored.ultimo_uso = datetime.utcnow()
    db.session.commit()
    login_user(stored.user)
    return jsonify({'success': True, 'redirect': url_for('main.dashboard')})


@main.route('/api/passkey/<int:credential_id>/delete', methods=['POST'])
@login_required
def passkey_delete(credential_id):
    credential = PasskeyCredential.query.filter_by(id=credential_id, user_id=current_user.id).first_or_404()
    db.session.delete(credential)
    db.session.commit()
    return jsonify({'success': True})

@main.route('/logout')
@login_required
def logout():
    current_user.ultima_actividad = None
    db.session.commit()
    logout_user()
    flash('Has cerrado sesión.', 'info')
    return redirect(url_for('main.login'))

@main.route('/submit_evaluacion', methods=['POST'])
@login_required
def submit_evaluacion():
    try:
        # Obtener el ID del usuario que será evaluado
        evaluado_id = request.form.get('evaluado_id')
        
        if not evaluado_id:
            flash("Usuario no seleccionado", "danger")
            return redirect(request.referrer)
        
        evaluado_id = int(evaluado_id)
        
        # Validar permisos - verificar en la BD
        permiso = PermisosEvaluacion.query.filter_by(
            evaluador_id=current_user.id,
            evaluado_id=evaluado_id
        ).first()
        
        if not permiso:
            flash("No tienes permiso para evaluar a este usuario", "danger")
            return redirect(request.referrer)
        
        # Verificar si ya existe una evaluación del mismo evaluador al mismo evaluado
        evaluacion_existente = EvaluacionDesempeno.query.filter_by(
            user_id=evaluado_id,
            evaluador_id=current_user.id
        ).first()
        
        if evaluacion_existente:
            flash("Ya has evaluado a esta persona. No puedes hacer una evaluación duplicada.", "warning")
            return redirect(request.referrer)
        
        # Recibimos los datos del formulario
        nombre = request.form.get('nombre')
        fecha_str = request.form.get('fecha')
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        area = request.form.get('area') or ""  # área es opcional
        estacion = request.form.get('estacion') or ""  # estación es opcional
        nomina = request.form.get('nomina')
        puesto = request.form.get('puesto')
        
        # Almacenamos las respuestas del formulario en un diccionario
        # Estructura genérica que soporta evaluaciones (subteniente, teniente y coordinador)
        respuestas = {
            # Para evaluación de Subteniente
            'comunicacion': {
                'claridad': request.form.get('comunicacion_claridad') or None,
                'escucha': request.form.get('comunicacion_escucha') or None,
                'equipo': request.form.get('comunicacion_equipo') or None,
                'conflictos': request.form.get('comunicacion_conflictos') or None
            },
            # Para evaluación de Teniente - Conocimientos Bomberiles
            'conocimientos': {
                'dominio_tecnico': request.form.get('conocimientos_dominio_tecnico') or None,
                'uso_equipo': request.form.get('conocimientos_uso_equipo') or None,
                'normatividad': request.form.get('conocimientos_normatividad') or None,
                'toma_decisiones': request.form.get('conocimientos_toma_decisiones') or None
            },
            # Para evaluación de Coordinador - Dirección de Guardias
            'direccion': {
                'direccion_guardias': request.form.get('direccion_guardias') or None,
                'toma_decisiones': request.form.get('direccion_toma_decisiones') or None,
                'motivacion': request.form.get('direccion_motivacion') or None,
                'normatividad': request.form.get('direccion_normatividad') or None
            },
            # Para evaluación de Coordinador - Gestión de Estaciones
            'gestion': {
                'organizacion': request.form.get('gestion_organizacion') or None,
                'protocolos': request.form.get('gestion_protocolos') or None,
                'supervisiona_mantenimiento': request.form.get('gestion_supervisiona_mantenimiento') or None,
                'adaptabilidad': request.form.get('gestion_adaptabilidad') or None
            },
            # Habilidades Blandas
            'habilidades_blandas': {
                'trabajo_equipo': request.form.get('blandas_trabajo_equipo') or None,
                'empatia': request.form.get('blandas_empatia') or None,
                'adaptabilidad': request.form.get('blandas_adaptabilidad') or None,
                'responsabilidad': request.form.get('blandas_responsabilidad') or None,
                'liderazgo': request.form.get('blandas_liderazgo') or None,
                'resolucion_conflictos': request.form.get('blandas_resolucion_conflictos') or None
            },
            # Comunicación con Subordinados (para Teniente y Coordinador)
            'comunicacion_subordinados': {
                'claridad_instrucciones': request.form.get('comunicacion_claridad_instrucciones') or None,
                'retroalimentacion': request.form.get('comunicacion_retroalimentacion') or None,
                'escucha_activa': request.form.get('comunicacion_escucha_activa') or None,
                'situaciones_criticas': request.form.get('comunicacion_situaciones_criticas') or None
            },
            # Disciplina
            'disciplina': {
                'puntualidad': request.form.get('disciplina_puntualidad') or None,
                'normas': request.form.get('disciplina_normas') or None,
                'presentacion': request.form.get('disciplina_presentacion') or None,
                'recursos': request.form.get('disciplina_recursos') or None,
                'imparcialidad': request.form.get('disciplina_imparcialidad') or None,
                'prevencion': request.form.get('disciplina_prevencion') or None,
                'documentacion': request.form.get('disciplina_documentacion') or None,
                'autoridad_moral': request.form.get('disciplina_autoridad_moral') or None
            },
            # Orden Cerrado
            'orden_cerrado': {
                'dominio': request.form.get('orden_cerrado_dominio') or None,
                'coordinacion': request.form.get('orden_cerrado_coordinacion') or None,
                'atencion': request.form.get('orden_cerrado_atencion') or None,
                'postura': request.form.get('orden_cerrado_postura') or None,
                'dominio_tecnicas': request.form.get('orden_cerrado_dominio_tecnicas') or None,
                'coordinacion_grupal': request.form.get('orden_cerrado_coordinacion_grupal') or None,
                'correccion_errores': request.form.get('orden_cerrado_correccion_errores') or None,
                'presencia_porte': request.form.get('orden_cerrado_presencia_porte') or None
            }
        }

        observaciones = request.form.get('observaciones')

        # Crear un objeto de la evaluación
        evaluacion = EvaluacionDesempeno(
            user_id=evaluado_id,
            evaluador_id=current_user.id,
            nombre=nombre,
            fecha=fecha,
            area=area,
            estacion=estacion,
            nomina=nomina,
            puesto=puesto,
            respuestas=respuestas,
            evaluacion_general=observaciones,
            comentario=observaciones
        )

        # Guardar la evaluación en la base de datos
        db.session.add(evaluacion)
        db.session.commit()

        flash("Evaluación enviada exitosamente", "success")  # Mostrar un mensaje de éxito

        # Redirigir al dashboard
        return redirect(url_for('main.dashboard'))
    
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error en submit_evaluacion: {str(e)}")
        flash(f"Error al guardar la evaluación: {str(e)}", "danger")
        return redirect(request.referrer)

@main.route('/resultados_evaluaciones')
@login_required
def resultados_evaluaciones():
    try:
        rol_seleccionado = request.args.get('rol', 'BOMBERO ESPECIALIZADO').strip()
        turno_seleccionado = request.args.get('turno', 'GENERAL').strip()
        
        logging.info(f"[DEBUG] Parámetros recibidos: rol='{rol_seleccionado}', turno='{turno_seleccionado}'")
        
        # Obtener todos los usuarios del rol seleccionado
        if turno_seleccionado == 'GENERAL':
            # Si es GENERAL, obtener todos los usuarios del rol sin filtrar por turno
            logging.info(f"[DEBUG] Filtrando GENERAL - obteniendo todos los usuarios con puesto='{rol_seleccionado}'")
            usuarios = User.query.filter_by(puesto=rol_seleccionado).all()
        else:
            # Si es un turno específico, filtrar por puesto y turno
            logging.info(f"[DEBUG] Filtrando turno específico - puesto='{rol_seleccionado}', turno='{turno_seleccionado}'")
            usuarios = User.query.filter_by(puesto=rol_seleccionado, turno=turno_seleccionado).all()
        
        usuario_ids = [u.id for u in usuarios]
        logging.info(f"[DEBUG] Usuarios encontrados: {len(usuarios)} con IDs: {usuario_ids[:10]}...")  # Mostrar primeros 10
        
        # Obtener todas las evaluaciones de estos usuarios
        if usuario_ids:
            evaluaciones = EvaluacionDesempeno.query.filter(
                EvaluacionDesempeno.user_id.in_(usuario_ids)
            ).all()
        else:
            evaluaciones = []
        
        # Definir categorías por rol evaluado
        categorias_por_rol = {
            'BOMBERO ESPECIALIZADO': {
                'comunicacion': {'titulo': 'Comunicación', 'valores': []},
                'habilidades_blandas': {'titulo': 'Habilidades Blandas', 'valores': []},
                'disciplina': {'titulo': 'Disciplina', 'valores': []},
                'orden_cerrado': {'titulo': 'Orden Cerrado', 'valores': []}
            },
            'SUBTENIENTE': {
                'conocimientos': {'titulo': 'Conocimientos Bomberiles', 'valores': []},
                'habilidades_blandas': {'titulo': 'Habilidades Blandas', 'valores': []},
                'comunicacion_subordinados': {'titulo': 'Comunicación con Subordinados', 'valores': []},
                'disciplina': {'titulo': 'Disciplina', 'valores': []},
                'orden_cerrado': {'titulo': 'Orden Cerrado', 'valores': []}
            },
            'TENIENTE': {
                'direccion': {'titulo': 'Dirección de Guardias y Turnos', 'valores': []},
                'gestion': {'titulo': 'Gestión de Estaciones', 'valores': []},
                'comunicacion_subordinados': {'titulo': 'Comunicación con Subordinados', 'valores': []},
                'disciplina': {'titulo': 'Disciplina', 'valores': []},
                'orden_cerrado': {'titulo': 'Orden Cerrado', 'valores': []}
            }
        }
        
        # Obtener categorías para el rol seleccionado
        categorias = categorias_por_rol.get(rol_seleccionado, categorias_por_rol['BOMBERO ESPECIALIZADO'])
        
        logging.info(f"Consultando evaluaciones para rol: {rol_seleccionado}, turno: {turno_seleccionado}")
        logging.info(f"Total de evaluaciones encontradas: {len(evaluaciones)}")
        
        for evaluacion in evaluaciones:
            if not evaluacion or not evaluacion.respuestas:
                continue
            
            respuestas = evaluacion.respuestas
            
            # Si respuestas es string (JSON encoded), deserializarlo
            if isinstance(respuestas, str):
                try:
                    respuestas = pyjson.loads(respuestas)
                except pyjson.JSONDecodeError:
                    logging.error(f"Error deserializando JSON en evaluación {evaluacion.id}")
                    continue
            
            if not isinstance(respuestas, dict):
                logging.warning(f"Respuestas no es dict: {type(respuestas)}")
                continue
            
            for categoria, datos in categorias.items():
                if categoria not in respuestas:
                    continue
                
                categoria_data = respuestas[categoria]
                if not isinstance(categoria_data, dict):
                    continue
                
                try:
                    # Convertir valores a int y filtrar los que no sean válidos
                    valores = []
                    for v in categoria_data.values():
                        if v:  # Skip None y empty strings
                            try:
                                valores.append(int(v))
                            except (ValueError, TypeError):
                                # Skip values that can't be converted to int
                                pass
                    
                    if valores:
                        promedio = sum(valores) / len(valores)
                        datos['valores'].append(promedio)
                except Exception as e:
                    logging.error(f"Error processing categoria {categoria}: {str(e)}")
                    continue
        
        # Calcular promedios finales y contar evaluaciones
        datos_grafica = []
        for categoria, datos in categorias.items():
            if datos['valores']:
                promedio_final = sum(datos['valores']) / len(datos['valores'])
            else:
                promedio_final = 0
            
            datos_grafica.append({
                'nombre': datos['titulo'],
                'promedio': round(promedio_final, 2),
                'cantidad_evaluaciones': len(datos['valores'])
            })
        
        return render_template(
            'resultados_evaluaciones.html',
            datos_grafica=datos_grafica,
            rol_seleccionado=rol_seleccionado,
            turno_seleccionado=turno_seleccionado,
            total_evaluaciones=len(evaluaciones),
            roles_disponibles=['BOMBERO ESPECIALIZADO', 'SUBTENIENTE', 'TENIENTE'],
            turnos_disponibles=['GENERAL', 'A', 'B', 'C']
        )
    except Exception as e:
        logging.error(f"Error en resultados_evaluaciones: {str(e)}")
        flash(f"Error al cargar los resultados: {str(e)}", "danger")
        return redirect(url_for('main.dashboard'))


@main.route('/descargar_resultados_excel', methods=['GET'])
@login_required
def descargar_resultados_excel():
    try:
        rol_seleccionado = request.args.get('rol', 'BOMBERO ESPECIALIZADO')
        turno_seleccionado = request.args.get('turno', 'GENERAL')
        
        # Obtener usuarios del rol seleccionado
        if turno_seleccionado == 'GENERAL':
            usuarios = User.query.filter_by(puesto=rol_seleccionado).all()
        else:
            usuarios = User.query.filter_by(puesto=rol_seleccionado, turno=turno_seleccionado).all()
        
        usuario_ids = [u.id for u in usuarios]
        
        # Obtener evaluaciones
        if usuario_ids:
            evaluaciones = EvaluacionDesempeno.query.filter(
                EvaluacionDesempeno.user_id.in_(usuario_ids)
            ).all()
        else:
            evaluaciones = []
        
        # Definir categorías por rol
        categorias_por_rol = {
            'BOMBERO ESPECIALIZADO': {
                'comunicacion': 'Comunicación',
                'habilidades_blandas': 'Habilidades Blandas',
                'disciplina': 'Disciplina',
                'orden_cerrado': 'Orden Cerrado'
            },
            'SUBTENIENTE': {
                'conocimientos': 'Conocimientos Bomberiles',
                'habilidades_blandas': 'Habilidades Blandas',
                'comunicacion_subordinados': 'Comunicación con Subordinados',
                'disciplina': 'Disciplina',
                'orden_cerrado': 'Orden Cerrado'
            },
            'TENIENTE': {
                'direccion': 'Dirección de Guardias y Turnos',
                'gestion': 'Gestión de Estaciones',
                'comunicacion_subordinados': 'Comunicación con Subordinados',
                'disciplina': 'Disciplina',
                'orden_cerrado': 'Orden Cerrado'
            }
        }
        
        categorias = categorias_por_rol.get(rol_seleccionado, categorias_por_rol['BOMBERO ESPECIALIZADO'])
        
        # Procesar datos de evaluaciones - DATOS AGREGADOS y DETALLADOS
        datos_categorias_agregados = {cat: [] for cat in categorias.keys()}
        datos_usuarios = {}  # {usuario_id: {categoria: [valores]}}
        
        for evaluacion in evaluaciones:
            if not evaluacion or not evaluacion.respuestas:
                continue
            
            usuario = User.query.get(evaluacion.user_id)
            evaluador = User.query.get(evaluacion.evaluador_id)
            if not usuario:
                continue
            
            respuestas = evaluacion.respuestas
            
            if isinstance(respuestas, str):
                try:
                    respuestas = pyjson.loads(respuestas)
                except:
                    continue
            
            if not isinstance(respuestas, dict):
                continue
            
            # Inicializar datos del usuario si no existen
            if evaluacion.user_id not in datos_usuarios:
                datos_usuarios[evaluacion.user_id] = {
                    'nombre': usuario.nombre,
                    'username': usuario.username,
                    'puesto': usuario.puesto,
                    'turno': usuario.turno or 'N/A',
                    'estacion': usuario.estacion or 'N/A',
                    'comentarios': [],
                    'categorias': {cat: [] for cat in categorias.keys()}
                }
            
            comentario_evaluacion = evaluacion.comentario.strip() if evaluacion.comentario else ''
            if comentario_evaluacion:
                datos_usuarios[evaluacion.user_id]['comentarios'].append(comentario_evaluacion)
            
            # Procesar cada categoría
            promedios_categorias = []
            for categoria in categorias.keys():
                if categoria not in respuestas:
                    continue
                
                categoria_data = respuestas[categoria]
                if not isinstance(categoria_data, dict):
                    continue
                
                valores = []
                for v in categoria_data.values():
                    if v:
                        try:
                            valores.append(int(v))
                        except (ValueError, TypeError):
                            pass
                
                if valores:
                    promedio = sum(valores) / len(valores)
                    promedios_categorias.append(promedio)
                    datos_categorias_agregados[categoria].append(promedio)
                    datos_usuarios[evaluacion.user_id]['categorias'][categoria].append(promedio)
            
            # Calcular calificación general como promedio de todos los promedios
            if promedios_categorias:
                calificacion_general = sum(promedios_categorias) / len(promedios_categorias)
                # No se utiliza al nivel usuario directo, pero sí permite conservar el cálculo si se necesita.
        
        # Crear Excel con dos hojas
        wb = openpyxl.Workbook()
        
        # Estilos
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        user_header_fill = PatternFill(start_color="0891b2", end_color="0891b2", fill_type="solid")
        user_header_font = Font(bold=True, color="FFFFFF", size=11)
        excellent_fill = PatternFill(start_color="86efac", end_color="86efac", fill_type="solid")  # Verde
        good_fill = PatternFill(start_color="fbbf24", end_color="fbbf24", fill_type="solid")  # Amarillo
        poor_fill = PatternFill(start_color="f87171", end_color="f87171", fill_type="solid")  # Rojo
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ===== HOJA 1: RESUMEN =====
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        ws_resumen['A1'] = "RESULTADOS DE EVALUACIONES DEL DESEMPEÑO"
        ws_resumen['A1'].font = Font(bold=True, size=14, color="1e293b")
        ws_resumen.merge_cells('A1:C1')
        
        ws_resumen['A2'] = f"Puesto: {rol_seleccionado}"
        ws_resumen['B2'] = f"Turno: {turno_seleccionado}"
        ws_resumen['C2'] = f"Total de Evaluaciones: {len(evaluaciones)}"
        
        # Tabla de resumen
        row = 4
        ws_resumen[f'A{row}'] = "Categoría"
        ws_resumen[f'B{row}'] = "Promedio (de 5)"
        ws_resumen[f'C{row}'] = "Cantidad de Evaluaciones"
        
        for col in ['A', 'B', 'C']:
            ws_resumen[f'{col}{row}'].fill = header_fill
            ws_resumen[f'{col}{row}'].font = header_font
            ws_resumen[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_resumen[f'{col}{row}'].border = thin_border
        
        row += 1
        for categoria, titulo in categorias.items():
            valores = datos_categorias_agregados[categoria]
            promedio = sum(valores) / len(valores) if valores else 0
            cantidad = len(valores)
            
            ws_resumen[f'A{row}'] = titulo
            ws_resumen[f'B{row}'] = round(promedio, 2)
            ws_resumen[f'C{row}'] = cantidad
            
            for col in ['A', 'B', 'C']:
                ws_resumen[f'{col}{row}'].border = thin_border
                ws_resumen[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
        
        ws_resumen.column_dimensions['A'].width = 35
        ws_resumen.column_dimensions['B'].width = 20
        ws_resumen.column_dimensions['C'].width = 25
        
        # ===== HOJA 2: DETALLADO POR USUARIO =====
        ws_detallado = wb.create_sheet("Detallado por Usuario")
        
        # Encabezados
        ws_detallado['A1'] = "EVALUACIONES DETALLADAS POR USUARIO"
        ws_detallado['A1'].font = Font(bold=True, size=14, color="1e293b")
        
        row = 3
        ws_detallado[f'A{row}'] = "Nómina"
        ws_detallado[f'B{row}'] = "Nombre del Evaluado"
        ws_detallado[f'C{row}'] = "Puesto"
        ws_detallado[f'D{row}'] = "Turno"
        ws_detallado[f'E{row}'] = "Comentarios"
        
        # Agregar columnas para cada categoría
        col_idx = 6
        categoria_cols = {}
        for categoria, titulo in categorias.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws_detallado[f'{col_letter}{row}'] = titulo
            categoria_cols[categoria] = col_letter
            col_idx += 1
        last_header_col = openpyxl.utils.get_column_letter(col_idx - 1)
        ws_detallado.merge_cells(f'A1:{last_header_col}1')
        
        # Estilos para encabezado
        for col in range(1, col_idx):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_detallado[f'{col_letter}{row}'].fill = user_header_fill
            ws_detallado[f'{col_letter}{row}'].font = user_header_font
            ws_detallado[f'{col_letter}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws_detallado[f'{col_letter}{row}'].border = thin_border
        
        # Llenar datos de usuarios
        row += 1
        for user_id in sorted(datos_usuarios.keys()):
            user_data = datos_usuarios[user_id]
            
            ws_detallado[f'A{row}'] = user_data['username']
            ws_detallado[f'B{row}'] = user_data['nombre']
            ws_detallado[f'C{row}'] = user_data['puesto']
            ws_detallado[f'D{row}'] = user_data['turno']
            comentarios_usuario = user_data['comentarios']
            ws_detallado[f'E{row}'] = '\n---\n'.join(comentarios_usuario) if comentarios_usuario else 'Sin comentarios'
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_detallado[f'{col}{row}'].border = thin_border
                ws_detallado[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Llenar promedios por categoría
            for categoria, col_letter in categoria_cols.items():
                valores = user_data['categorias'][categoria]
                if valores:
                    promedio = sum(valores) / len(valores)
                    ws_detallado[f'{col_letter}{row}'] = round(promedio, 2)
                    
                    # Aplicar color según el promedio
                    if promedio >= 4.5:
                        fill_color = excellent_fill
                    elif promedio >= 3.5:
                        fill_color = good_fill
                    else:
                        fill_color = poor_fill
                    
                    ws_detallado[f'{col_letter}{row}'].fill = fill_color
                else:
                    ws_detallado[f'{col_letter}{row}'] = "N/A"
                
                ws_detallado[f'{col_letter}{row}'].border = thin_border
                ws_detallado[f'{col_letter}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
        
        ws_detallado.column_dimensions['A'].width = 18  
        ws_detallado.column_dimensions['B'].width = 30  
        ws_detallado.column_dimensions['C'].width = 20
        ws_detallado.column_dimensions['D'].width = 12
        ws_detallado.column_dimensions['E'].width = 40
        for categoria, col_letter in categoria_cols.items():
            ws_detallado.column_dimensions[col_letter].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        fecha = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"Evaluaciones_{rol_seleccionado.replace(' ', '_')}_{turno_seleccionado}_{fecha}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logging.error(f"Error al descargar resultados en Excel: {str(e)}")
        flash(f"Error al descargar los resultados: {str(e)}", "danger")
        return redirect(url_for('main.resultados_evaluaciones'))


@main.route('/api/evaluaciones_disponibles', methods=['GET'])
@login_required
def api_evaluaciones_disponibles():
    """API que retorna todas las evaluaciones disponibles para búsqueda"""
    try:
        evaluaciones = EvaluacionDesempeno.query.all()
        evaluaciones_list = []
        
        for eval in evaluaciones:
            usuario = User.query.get(eval.user_id)
            if usuario:
                evaluaciones_list.append({
                    'eval_id': eval.id,
                    'user_id': eval.user_id,
                    'nombre': usuario.nombre,
                    'nomina': usuario.username,
                    'puesto': usuario.puesto,
                    'turno': usuario.turno or 'N/A',
                    'estacion': usuario.estacion or 'N/A',
                    'fecha': eval.fecha.strftime('%d/%m/%Y') if eval.fecha else 'N/A'
                })
        
        return jsonify({'evaluaciones': evaluaciones_list})
    except Exception as e:
        logging.error(f"Error en API evaluaciones: {str(e)}")
        return jsonify({'error': str(e)}), 500


@main.route('/api/evaluacion/<int:eval_id>', methods=['GET'])
@login_required
def api_evaluacion_detalle(eval_id):
    """API que retorna los datos de una evaluación específica"""
    try:
        evaluacion = EvaluacionDesempeno.query.get(eval_id)
        if not evaluacion:
            return jsonify({'error': 'Evaluación no encontrada'}), 404
        
        usuario_evaluado = User.query.get(evaluacion.user_id)
        usuario_evaluador = User.query.get(evaluacion.evaluador_id)
        
        if not usuario_evaluado:
            return jsonify({'error': 'Usuario evaluado no encontrado'}), 404
        
        respuestas = evaluacion.respuestas
        temas_calificaciones = {}
        
        for tema, preguntas_dict in respuestas.items():
            valores = [int(v) for v in preguntas_dict.values() if v]
            if valores:
                promedio = round(sum(valores) / len(valores), 2)
                tema_nombres = {
                    'comunicacion': 'Comunicación',
                    'habilidades_blandas': 'Habilidades Blandas',
                    'disciplina': 'Disciplina',
                    'orden_cerrado': 'Orden Cerrado',
                    'conocimientos': 'Conocimientos Bomberiles',
                    'comunicacion_subordinados': 'Comunicación con Subordinados',
                    'liderazgo': 'Liderazgo'
                }
                tema_nombre = tema_nombres.get(tema, tema)
                temas_calificaciones[tema_nombre] = promedio
        
        return jsonify({
            'evaluacion': {
                'id': evaluacion.id,
                'fecha': evaluacion.fecha.strftime('%d/%m/%Y') if evaluacion.fecha else 'N/A',
                'comentario': evaluacion.comentario or '',
            },
            'usuario_evaluado': {
                'nombre': usuario_evaluado.nombre,
                'puesto': usuario_evaluado.puesto,
                'turno': usuario_evaluado.turno or 'N/A',
                'estacion': usuario_evaluado.estacion or 'N/A',
                'username': usuario_evaluado.username
            },
            'usuario_evaluador': {
                'nombre': usuario_evaluador.nombre if usuario_evaluador else 'N/A',
            },
            'temas_calificaciones': temas_calificaciones
        })
    except Exception as e:
        logging.error(f"Error en API evaluacion detalle: {str(e)}")
        return jsonify({'error': str(e)}), 500


@main.route('/submit_form', methods=['POST'])
@login_required
def submit_form():
    nombre = request.form.get('nombre')
    email = request.form.get('email')
    mensaje = request.form.get('mensaje')
    flash(f'Mensaje recibido de {nombre}', 'success')
    return redirect(url_for('main.dashboard'))

@main.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename, as_attachment=True)


@main.before_request
def verificar_inactividad():
    session.permanent = True  
    session.modified = True  

    if 'ultima_actividad' in session:
        ultima_actividad_str = session.get('ultima_actividad')

        if isinstance(ultima_actividad_str, str): 
            try:
                ultima_actividad = datetime.fromisoformat(ultima_actividad_str)  
            except ValueError:
                ultima_actividad = datetime.now()  
        else:
            ultima_actividad = datetime.now()  
        
        tiempo_inactivo = datetime.now() - ultima_actividad
        if tiempo_inactivo > timedelta(hours=2):
            logout_user()
            session.clear()
            flash('Tu sesión ha expirado por inactividad.', 'warning')
            return redirect(url_for('main.login'))
    session['ultima_actividad'] = datetime.now().isoformat()  

@main.route('/check_vacation_status', methods=['GET'])
@login_required
def check_vacation_status():
    vacation = VacationRequest.query.filter_by(user_id=current_user.id).first()
    if vacation:
        return jsonify({
            "sent": True,
            "selected_date": vacation.selected_date.strftime('%Y-%m-%d'),
            "assigned_date": vacation.assigned_date.strftime('%Y-%m-%d')
        })
    return jsonify({"sent": False})

@main.route('/save_vacation_date', methods=['POST'])
@login_required
def save_vacation_date():
    if VacationRequest.query.filter_by(user_id=current_user.username).first():
        return jsonify({"error": "Ya has enviado una fecha. Solo puedes enviarla una vez."}), 400
    data = request.get_json()
    selected_date_str = data.get('selected_date')
    assigned_date_str = data.get('assigned_date')
    if not selected_date_str or not assigned_date_str:
        return jsonify({"error": "Faltan datos de fecha."}), 400
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        assigned_date = datetime.strptime(assigned_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"error": "Formato de fecha incorrecto."}), 400
    new_request = VacationRequest(
        user_id=current_user.username,
        selected_date=selected_date,
        assigned_date=assigned_date
    )
    db.session.add(new_request)
    db.session.commit()
    return jsonify({"message": "Fecha guardada exitosamente.", "redirect": "dashboard"})

@main.route('/ver-excel')
def ver_excel():
    excel_path = "static/documentos/calendario.xlsx"
    df = pd.read_excel(excel_path)
    excel_html = df.to_html(classes="table table-striped", index=False)
    return render_template("ver_excel.html", excel_html=excel_html)

@main.route('/profile')
@login_required
def profile():
    return render_template('profile.html')

@main.route('/export_competencia_excel')
@login_required
def export_competencia_excel():
    # Obtener todos los registros de la tabla RegistroCompetencia
    registros = RegistroCompetencia.query.all()

    # Crear lista de diccionarios para convertir a DataFrame
    datos = []
    for r in registros:
        datos.append({
            "Número Competidor": r.numero_competidor,
            "User ID": r.user_id,
            "Nombre": r.nombre,
            "Nómina": r.nomina,
            "Turno": r.turno,
            "Categoría": r.categoria,
            "Niños": r.ninos,
            "Adultos": r.adultos,
            "Correo": r.correo,
        })

    # Crear DataFrame
    df = pd.DataFrame(datos)

    # Opcional: reordenar columnas si quieres
    columnas_orden = ["Número Competidor", "User ID", "Nombre", "Nómina", "Turno", "Categoría", "Niños", "Adultos", "Correo"]
    df = df[columnas_orden]

    # Crear archivo Excel en memoria
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Registro Competencia')

        # Formato: ajustar ancho columnas
        workbook = writer.book
        worksheet = writer.sheets['Registro Competencia']

        for i, col in enumerate(df.columns):
            # Ajustar ancho con base en la longitud del contenido y el nombre de la columna
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(col)
            ) + 2  # espacio extra
            worksheet.set_column(i, i, max_len)

    output.seek(0)

    # Enviar archivo para descargar
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='registro_competencia.xlsx',
                     as_attachment=True)

@main.route('/gestor_contenido', defaults={'folder_id': None})
@main.route('/gestor_contenido/<int:folder_id>')
@login_required
def gestor_contenido(folder_id):
    if folder_id:
        current_folder = Folder.query.get(folder_id)
        files = File.query.filter_by(folder_id=folder_id).all()
        folders = Folder.query.filter_by(parent_id=folder_id).all()
    else:
        current_folder = None
        files = File.query.filter_by(folder_id=None).all()
        folders = Folder.query.filter_by(parent_id=None).all() 
    return render_template('gestor_contenido.html', folders=folders, current_folder=current_folder, files=files)
@main.route('/upload_file', methods=['POST'])
@login_required
def upload_file():
    file = request.files.get('file')
    folder_id = request.form.get('folder_id')
    if file and folder_id:
        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        new_file = File(filename=filename, folder_id=folder_id)
        db.session.add(new_file)
        db.session.commit()

        return jsonify({'success': True, 'filename': filename})
    return jsonify({'success': False, 'message': 'Faltan datos'}), 400


@main.route('/add_folder', methods=['POST'])
@login_required
def add_folder():
    folder_name = request.form.get('folder_name')
    parent_id = request.form.get('parent_id') or None 
    if not folder_name:
        return jsonify({'success': False, 'message': 'El nombre de la carpeta es obligatorio.'}), 400
    try:
        new_folder = Folder(name=folder_name, parent_id=parent_id)
        db.session.add(new_folder)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Carpeta creada correctamente.'}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error al crear la carpeta: {str(e)}'}), 500


@main.route('/rename_folder/<int:folder_id>', methods=['POST'])
@login_required
def rename_folder(folder_id):
    folder = Folder.query.get(folder_id)
    if folder:
        data = request.get_json()
        folder.name = data.get('name')
        db.session.commit()
        return jsonify({'success': True, 'message': 'Carpeta renombrada correctamente.'})
    return jsonify({'success': False, 'message': 'Carpeta no encontrada.'}), 404

@main.route('/move_folder/<int:folder_id>', methods=['POST'])
@login_required
def move_folder(folder_id):
    folder = Folder.query.get(folder_id)
    if folder:
        data = request.get_json()
        folder.parent_id = data.get('parent_id')
        db.session.commit()
        return jsonify({'success': True, 'message': 'Carpeta movida correctamente.'})
    return jsonify({'success': False, 'message': 'Carpeta no encontrada.'}), 404

@main.route('/get_all_folders', methods=['GET'])
@login_required
def get_all_folders():
    folders = Folder.query.all() 
    folder_list = [{"id": folder.id, "name": folder.name, "parent_id": folder.parent_id} for folder in folders]
    return jsonify({"success": True, "folders": folder_list})

@main.route('/rename_file/<int:file_id>', methods=['POST'])
@login_required
def rename_file(file_id):
    file = File.query.get(file_id)
    if file:
        data = request.get_json()
        new_name = data.get('new_name')
        if new_name:
            file.filename = new_name
            db.session.commit()
            return jsonify({'success': True, 'message': 'Archivo renombrado correctamente.'})
    return jsonify({'success': False, 'message': 'Archivo no encontrado.'}), 404

@main.route('/add_aviso', methods=['POST'])
@login_required
def add_aviso():
    data = request.get_json()
    descripcion = data.get('descripcion')
    fecha_caducidad = data.get('fecha_caducidad')

    if descripcion:
        aviso = Aviso(
            descripcion=descripcion,
            fecha_caducidad=datetime.strptime(fecha_caducidad, '%Y-%m-%d').date() if fecha_caducidad else None,
            user_id=current_user.id
        )
        db.session.add(aviso)
        db.session.commit()

        return jsonify({
            "success": True,
            "aviso": {
                "descripcion": aviso.descripcion,
                "fecha_caducidad": aviso.fecha_caducidad.strftime('%Y-%m-%d') if aviso.fecha_caducidad else "",
                "fecha_creacion": aviso.fecha_creacion.strftime('%Y-%m-%d %H:%M')
            }
        })
    return jsonify({"success": False}), 400

@main.route('/formulario')
@login_required
def formulario():
    return render_template('formulario.html')

@main.route('/formulario_datos', methods=['GET', 'POST'])
@login_required
def formulario_datos():
    if request.method == 'POST':
        return submit_formulario()
    return render_template('formulario.html')

@main.route('/api/buscar_usuario/<username>')
@login_required
def buscar_usuario(username):
    usuario = User.query.filter_by(username=username).first()
    if usuario:
        return jsonify({
            "success": True,
            "id": usuario.id,
            "nombre": usuario.nombre,
            "puesto": usuario.puesto,
            "turno": usuario.turno or "",
            'estacion': usuario.estacion,
            "nomina": usuario.username,
            "image_file": usuario.image_file or "default.jpg"
        })
    return jsonify({"success": False, "error": "Usuario no encontrado"}), 404
@main.route('/api/listar_usuarios')
@login_required
def listar_usuarios():
    filtros = {
        "COORDINADOR OPERATIVO": ["TENIENTE"],
        "SUBTENIENTE": ["BOMBERO ESPECIALIZADO", "BOMBERO HABILITADO"],
        "TENIENTE": ["SUBTENIENTE"],
    }

    query = User.query.filter(User.id != current_user.id)

    if current_user.puesto in filtros:
        query = query.filter(User.puesto.in_(filtros[current_user.puesto]))

    usuarios = query.all()

    lista = [
        {
            "username": u.username,
            "nombre": u.nombre,
            "turno": u.turno or ""
        }
        for u in usuarios
    ]

    return jsonify({"success": True, "usuarios": lista})

@main.route('/api/listar_usuarios_permitidos')
@login_required
def listar_usuarios_permitidos():
    """API que retorna usuarios que el evaluador actual tiene permiso de evaluar"""
    evaluador_id = current_user.id
    
    # Obtener usuarios permitidos desde PermisosEvaluacion
    permisos = PermisosEvaluacion.query.filter_by(evaluador_id=evaluador_id).all()
    usuarios_permitidos_ids = [p.evaluado_id for p in permisos]
    
    if not usuarios_permitidos_ids:
        return jsonify({"success": True, "usuarios": []})
    
    usuarios = User.query.filter(User.id.in_(usuarios_permitidos_ids)).all()
    
    lista = [
        {
            "id": u.id,
            "username": u.username,
            "nombre": u.nombre,
            "turno": u.turno or "N/A",
            "puesto": u.puesto or "N/A"
        }
        for u in usuarios
    ]
    
    return jsonify({"success": True, "usuarios": lista})

@main.route('/api/listado_fotos')
@login_required
def listado_fotos():
    """Retorna lista de usuarios con URLs de fotos para reconocimiento facial."""
    usuarios = User.query.filter(
        ~db.func.lower(User.username).in_(USUARIOS_SIN_FUNCIONES)
    ).all()
    import os
    
    usuarios_info = []
    firma_partes = []
    for usuario in usuarios:
        foto_path = os.path.join(current_app.root_path, 'static', 'uploads', f'{usuario.username}.jpg')
        if os.path.exists(foto_path):
            mtime = int(os.path.getmtime(foto_path))
            usuarios_info.append({
                "username": usuario.username,
                "nombre": usuario.nombre,
                "foto_url": url_for('static', filename=f'uploads/{usuario.username}.jpg'),
                "mtime": mtime
            })
            firma_partes.append(f"{usuario.username}:{mtime}")
    
    firma = "|".join(sorted(firma_partes))
    return jsonify({"success": True, "usuarios": usuarios_info, "firma": firma})

def _face_cache_file_path():
    os.makedirs(current_app.instance_path, exist_ok=True)
    return os.path.join(current_app.instance_path, 'face_references_cache.json')


@main.route('/api/face_cache', methods=['GET', 'POST'])
@login_required
def face_cache():
    usuarios_resp = listado_fotos().get_json()
    current_firma = usuarios_resp.get('firma', '')

    cache_path = _face_cache_file_path()

    if request.method == 'GET':
        if not os.path.exists(cache_path):
            return jsonify({"success": True, "ready": False, "firma": current_firma, "items": []})

        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                payload = pyjson.load(f)
        except Exception:
            return jsonify({"success": True, "ready": False, "firma": current_firma, "items": []})

        ready = payload.get('firma') == current_firma and len(payload.get('items', [])) > 0
        return jsonify({
            "success": True,
            "ready": ready,
            "firma": current_firma,
            "items": payload.get('items', []) if ready else []
        })

    body = request.get_json(silent=True) or {}
    incoming_firma = body.get('firma', '')
    items = body.get('items', [])

    if incoming_firma != current_firma:
        return jsonify({"success": False, "error": "Firma desactualizada"}), 409

    if not isinstance(items, list) or len(items) == 0:
        return jsonify({"success": False, "error": "Sin descriptores para guardar"}), 400

    payload = {"firma": current_firma, "items": items}
    with open(cache_path, 'w', encoding='utf-8') as f:
        pyjson.dump(payload, f)

    return jsonify({"success": True})


@main.route('/api/facial_status')
def facial_status():
    """
    Retorna el estado del servicio de reconocimiento facial.
    """
    try:
        from app.facial_service import facial_service
        
        is_ready = (
            facial_service.initialized and 
            facial_service.embeddings_model is not None and
            len(facial_service.face_descriptors) > 0
        )
        
        return jsonify({
            "success": True,
            "ready": is_ready,
            "count": len(facial_service.face_descriptors)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "ready": False,
            "error": str(e)
        }), 500


@main.route('/api/recognize_face', methods=['POST'])
def recognize_face():
    """
    Reconoce un rostro en una imagen enviada como base64.
    Usa el servicio facial precargado en servidor.
    """
    return jsonify({
        "success": False,
        "error": "Endpoint sustituido por la verificacion privada en el dispositivo"
    }), 410

    try:
        from app.facial_service import facial_service
        
        if not facial_service.initialized or not facial_service.face_descriptors:
            return jsonify({
                "success": False,
                "error": "Servicio facial no disponible"
            }), 503
        
        body = request.get_json(silent=True) or {}
        image_base64 = body.get('image', '').strip()
        
        if not image_base64:
            return jsonify({
                "success": False,
                "error": "Imagen requerida"
            }), 400
        
        # Decodificar imagen base64
        try:
            import base64
            import cv2
            import numpy as np
            
            # Remover prefijo data:image si existe
            if ',' in image_base64:
                image_base64 = image_base64.split(',')[1]
            
            image_data = base64.b64decode(image_base64)
            nparr = np.frombuffer(image_data, np.uint8)
            image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            if image is None:
                return jsonify({
                    "success": False,
                    "error": "Imagen inválida"
                }), 400
            
            # Reconocer rostro
            username, nombre, confidence = facial_service.recognize_face(image, threshold=0.5)
            
            if username:
                return jsonify({
                    "success": True,
                    "username": username,
                    "nombre": nombre,
                    "confidence": float(confidence)
                })
            else:
                return jsonify({
                    "success": False,
                    "error": "No se pudo identificar el rostro"
                }), 401
        
        except Exception as e:
            logger = logging.getLogger(__name__)
            logger.error(f"[Facial] Error decodificando imagen: {e}")
            return jsonify({
                "success": False,
                "error": "Error procesando imagen"
            }), 400
    
    except ImportError:
        return jsonify({
            "success": False,
            "error": "Servicio facial no disponible"
        }), 503
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"[Facial] Error en recognize_face: {e}")
        return jsonify({
            "success": False,
            "error": "Error interno del servidor"
        }), 500


def _validar_recaptcha_facial(recaptcha_token):
    """Valida reCAPTCHA sin conservar autorizaciones reutilizables en la sesion."""
    if not recaptcha_token:
        return False, "Completa la verificacion de seguridad"

    secret_key = current_app.config.get('RECAPTCHA_PRIVATE_KEY')
    if not secret_key:
        return False, "Configuracion de captcha no disponible"

    try:
        verify_resp = requests.post(
            'https://www.google.com/recaptcha/api/siteverify',
            data={
                'secret': secret_key,
                'response': recaptcha_token,
                'remoteip': request.remote_addr
            },
            timeout=8
        )
        verify_data = verify_resp.json()
    except Exception:
        return False, "No se pudo validar captcha"

    if not verify_data.get('success'):
        return False, "Captcha invalido o expirado"
    return True, None


def _limite_intentos_facial():
    ahora = datetime.utcnow().timestamp()
    intentos = [
        float(valor) for valor in session.get('facial_attempts', [])
        if ahora - float(valor) < 300
    ]
    if len(intentos) >= 5:
        session['facial_attempts'] = intentos
        return False
    intentos.append(ahora)
    session['facial_attempts'] = intentos
    return True


@main.route('/api/facial/challenge', methods=['POST'])
def crear_reto_facial():
    return jsonify({'success': False, 'error': 'El acceso facial anterior fue sustituido por passkeys'}), 410

    body = request.get_json(silent=True) or {}
    username = (body.get('username') or '').strip()
    recaptcha_token = (body.get('recaptcha_token') or '').strip()

    if not username:
        return jsonify({"success": False, "error": "Usuario requerido"}), 400

    user = User.query.filter_by(username=username).first()
    if not user:
        return jsonify({"success": False, "error": "Usuario no encontrado"}), 404

    if user.username.lower() in USUARIOS_SIN_FUNCIONES:
        return jsonify({"success": False, "error": "Este usuario requiere acceso convencional"}), 403

    foto_relativa = f'uploads/{user.username}.jpg'
    foto_path = os.path.join(current_app.root_path, 'static', foto_relativa)
    if not os.path.isfile(foto_path):
        return jsonify({"success": False, "error": "Tu cuenta no tiene una foto de perfil registrada"}), 404

    if not _limite_intentos_facial():
        return jsonify({"success": False, "error": "Demasiados intentos. Espera cinco minutos"}), 429

    captcha_ok, captcha_error = _validar_recaptcha_facial(recaptcha_token)
    if not captcha_ok:
        return jsonify({"success": False, "error": captcha_error}), 403

    token = secrets.token_urlsafe(32)
    session['facial_challenge'] = {
        'username': user.username,
        'token': token,
        'expires_at': datetime.utcnow().timestamp() + 120
    }
    return jsonify({
        "success": True,
        "challenge": token,
        "photo_url": url_for('static', filename=foto_relativa),
        "nombre": user.nombre
    })


@main.route('/api/facial/complete', methods=['POST'])
def completar_login_facial():
    return jsonify({'success': False, 'error': 'El acceso facial anterior fue sustituido por passkeys'}), 410

    body = request.get_json(silent=True) or {}
    username = (body.get('username') or '').strip()
    token = (body.get('challenge') or '').strip()
    reto = session.pop('facial_challenge', None)

    if not reto or not token:
        return jsonify({"success": False, "error": "La verificacion facial expiro"}), 403
    if datetime.utcnow().timestamp() > float(reto.get('expires_at', 0)):
        return jsonify({"success": False, "error": "La verificacion facial expiro"}), 403
    if not secrets.compare_digest(token, str(reto.get('token', ''))):
        return jsonify({"success": False, "error": "Verificacion facial invalida"}), 403
    if username != reto.get('username'):
        return jsonify({"success": False, "error": "El usuario no coincide con el reto"}), 403

    user = User.query.filter_by(username=username).first()
    if not user or user.username.lower() in USUARIOS_SIN_FUNCIONES:
        return jsonify({"success": False, "error": "Usuario no autorizado"}), 403

    login_user(user)
    session.pop('facial_attempts', None)
    return jsonify({"success": True, "redirect": url_for('main.dashboard')})


@main.route('/api/login_facial', methods=['POST'])
def login_facial_obsoleto():
    """Impide que el endpoint anterior autentique confiando solo en un username."""
    return jsonify({
        "success": False,
        "error": "Actualiza la pagina para usar el nuevo acceso facial"
    }), 410


@main.route('/api/verify_facial_captcha', methods=['POST'])
def verify_facial_captcha():
    return jsonify({
        "success": False,
        "error": "Actualiza la pagina para usar el nuevo acceso facial"
    }), 410

    body = request.get_json(silent=True) or {}
    recaptcha_token = (body.get('recaptcha_token') or '').strip()

    if not recaptcha_token:
        return jsonify({"success": False, "error": "Captcha requerido"}), 400

    secret_key = current_app.config.get('RECAPTCHA_PRIVATE_KEY')
    if not secret_key:
        return jsonify({"success": False, "error": "Configuracion de captcha no disponible"}), 500

    try:
        verify_resp = requests.post(
            'https://www.google.com/recaptcha/api/siteverify',
            data={
                'secret': secret_key,
                'response': recaptcha_token,
                'remoteip': request.remote_addr
            },
            timeout=8
        )
        verify_data = verify_resp.json()
    except Exception:
        return jsonify({"success": False, "error": "No se pudo validar captcha"}), 502

    if not verify_data.get('success'):
        return jsonify({"success": False, "error": "Captcha invalido o expirado"}), 403

    session['facial_captcha_verified_at'] = datetime.utcnow().timestamp()
    return jsonify({"success": True})

@main.route('/evaluacion_del_desempeño')
@login_required
def evaluacion():

    if current_user.id in [7, 263, 63]:
        return render_template('evaluacion_personal.html', usuario_actual=current_user)

    elif current_user.puesto == "SUBTENIENTE":
        return render_template('evaluacion_subteniente.html', usuario_actual=current_user)

    elif current_user.puesto == "TENIENTE":
        return render_template('evaluacion_teniente.html', usuario_actual=current_user)

    elif current_user.puesto == "COORDINADOR OPERATIVO":
        return render_template('evaluacion_coordinador.html', usuario_actual=current_user)

    else:
        flash('No tienes permiso para acceder a evaluaciones', 'warning')
        return redirect(url_for('main.dashboard'))

@main.route('/evaluacion_personal')
@login_required
def evaluacion_personal():
    """Ruta para evaluación personal con permisos desde BD"""
    evaluador_id = current_user.id
    
    # Verificar que el usuario tiene permisos para evaluar
    permisos = PermisosEvaluacion.query.filter_by(evaluador_id=evaluador_id).all()
    
    if not permisos:
        flash('No tienes usuarios asignados para evaluar', 'warning')
        return redirect(url_for('main.dashboard'))
    
    return render_template('evaluacion_personal.html', usuario_actual=current_user)

@main.route('/findeaño12w')
@login_required
def fin_anio():
    respuesta_existente = AsistenciaFinAnio.query.filter_by(user_id=current_user.id).first()
    if respuesta_existente:
        return render_template('dashboard.html', ya_respondio=True, registro=respuesta_existente)
    else:
        return render_template('dia_bombero.html', ya_respondio=False)

@main.route('/competencia_bomb')
@login_required
def competencia_bomb():
    registro_existente = RegistroCompetencia.query.filter_by(user_id=current_user.id).first()

    if registro_existente:
        return render_template('dashboard.html', ya_respondio=True)
    return render_template('competencia_bomb.html')

@main.route('/scanner_asistencia')
@login_required
def scanner_asistencia():
    respuestas_confirmadas = Respuesta.query.filter_by(respondido=True).all()
    return render_template('scanner.html', respuestas_confirmadas=respuestas_confirmadas)

@main.route('/api/marcar_asistencia', methods=['POST'])
@login_required
def marcar_asistencia():
    data = request.get_json()
    user_id = data.get('user_id')
    
    if not user_id:
        return jsonify({"success": False, "error": "user_id no proporcionado"}), 400
    
    print(f"Recibido user_id: {user_id}")
    respuesta = Respuesta.query.filter_by(user_id=user_id, respondido=True).first()

    if not respuesta:
        print(f"No se encontró la respuesta para user_id: {user_id}")
        return jsonify({"success": False, "error": "Usuario no encontrado o no respondido"}), 400

    if not respuesta.asistio:
        respuesta.asistio = True
        db.session.commit()
        print(f"Asistencia confirmada para: {respuesta.user.nombre}")
        return jsonify({"success": True, "nombre": respuesta.user.nombre, "nuevo_registro": True}), 200
    else:
        print(f"El usuario ya estaba registrado como asistido para user_id: {user_id}")
        return jsonify({"success": True, "nombre": respuesta.user.nombre, "nuevo_registro": False}), 200

@main.route('/bombero', methods=['GET'])
@login_required
def formulario_bombero():
    """Muestra el formulario de confirmación para el Día del Bombero"""
    return render_template('formulario_bombero.html')

@main.route('/submit_bombero', methods=['POST'])
@login_required
def submit_bombero():
    nombre = current_user.nombre
    acompanante = request.form.get("lleva_acompanante")
    nombre_acompanante = request.form.get("nombre_acompanante")
    tipo_acompanante = request.form.get("tipo_acompanante")  # NUEVO CAMPO
    correo = request.form.get("correo")

    respuesta = Respuesta.query.filter_by(user_id=current_user.id).first()
    if not respuesta:
        respuesta = Respuesta(user_id=current_user.id)

    respuesta.nombre_acompanante = nombre_acompanante
    respuesta.tipo_acompanante = tipo_acompanante  # Guardar el tipo
    respuesta.correo = correo
    respuesta.respondido = True

    db.session.add(respuesta)
    db.session.commit()

    contenido_qr = f"{current_user.id}\nAsistencia confirmada:\n{nombre}"
    if acompanante == "sí":
        contenido_qr += f"\nAcompañante: {nombre_acompanante}"
        if tipo_acompanante:
            contenido_qr += f" ({tipo_acompanante})"
    contenido_qr += f"\nCorreo: {correo}"

    qr = qrcode.make(contenido_qr)
    filename = f"{nombre.replace(' ', '_')}_qr.png"
    carpeta_qr = os.path.join(os.getcwd(), 'app', 'qr_codes')
    os.makedirs(carpeta_qr, exist_ok=True)
    filepath = os.path.join(carpeta_qr, filename)
    qr.save(filepath)

    html_body = f"""
    <html>
        <body style="font-family: Arial, sans-serif; background-color: #f5f5f5; padding: 20px; text-align: center;">
            <div style="max-width: 600px; margin: auto; background-color: white; padding: 30px; border-radius: 10px; box-shadow: 0 4px 10px rgba(0,0,0,0.1);">
                <h2 style="color: #d32f2f;">🎉 Confirmación de Asistencia</h2>
                <p>Hola <strong>{nombre}</strong>,</p>
                <p>Gracias por confirmar tu asistencia al <strong>Día del Bombero</strong>.</p>
                <p><strong>Este es tu pase de entrada</strong>. Presenta este código QR al ingresar al evento:</p>
                <img src="cid:{filename}" style="margin-top: 20px; width: 250px; height: auto;" alt="QR Code" />
                <p style="margin-top: 20px; font-size: 0.9rem; color: #555;">⚠️ Sin este pase no podrás acceder al evento. Por favor, no lo pierdas.</p>
                <p style="margin-top: 30px;">¡Te esperamos! 👨‍🚒</p>
            </div>
        </body>
    </html>
    """

    msg = Message("🎫 Tu pase para el Día del Bombero",
                sender="cristian.rodriguez@bomberosdeleon.org",
                recipients=[correo])

    msg.body = f"Hola {nombre}, adjunto encontrarás tu código QR para el evento."
    msg.html = html_body

    with open(filepath, 'rb') as fp:
        msg.attach(filename, "image/png", fp.read(), headers={"Content-ID": f"<{filename}>"})

    mail.send(msg)

    return render_template('confirmation.html', nombre=nombre, correo=correo)

import io
import pandas as pd
from flask import send_file

@main.route('/descargar_registros_excel', methods=['GET'])
@login_required
def descargar_registros_excel():
    # Obtener todos los registros
    registros = Respuesta.query.all()

    # Preparar los datos para el DataFrame
    data = []
    for r in registros:
        tiene_acompanante = bool(r.nombre_acompanante and r.nombre_acompanante.strip())
        data.append({
            "Nombre": r.user.nombre if r.user else "Desconocido",
            "Correo": r.correo,
            "Acompañante": r.nombre_acompanante if tiene_acompanante else "",
            "Tipo Acompañante": r.tipo_acompanante if tiene_acompanante else "",
            "Con Acompañante": "Sí" if tiene_acompanante else "No"
        })

    # Crear el DataFrame
    df = pd.DataFrame(data)

    # Escribir a un archivo Excel en memoria
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Participantes', index=False)
        workbook  = writer.book
        worksheet = writer.sheets['Participantes']

        # Estilo para resaltar en amarillo
        highlight_format = workbook.add_format({'bg_color': '#FFFF00'})

        # Aplicar formato a filas con acompañante
        for i, row in df.iterrows():
            if row['Con Acompañante'] == 'Sí':
                worksheet.set_row(i + 1, None, highlight_format)  # +1 por encabezado

    output.seek(0)

    return send_file(output,
                     download_name="registros_bombero.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main.route('/submit_competencia', methods=['POST'])
@login_required
def submit_competencia():
    nombre = request.form.get('nombre')
    nomina = request.form.get('nomina')
    turno = request.form.get('turno')
    categoria = request.form.get('categoria')
    ninos = int(request.form.get('ninos', 0))
    adultos = int(request.form.get('adultos', 0))
    correo = request.form.get('correo')
    registro = RegistroCompetencia.query.filter_by(user_id=current_user.id).first()

    if not registro:
        ultimo_num = db.session.query(db.func.max(RegistroCompetencia.numero_competidor)).scalar()
        nuevo_num = 1 if ultimo_num is None else ultimo_num + 1

        registro = RegistroCompetencia(
            user_id=current_user.id,
            nombre=nombre,
            nomina=nomina,
            turno=turno,
            categoria=categoria,
            ninos=ninos,
            adultos=adultos,
            numero_competidor=nuevo_num,
            correo=correo
        )
        db.session.add(registro)
        db.session.commit()
    else:
        nuevo_num = registro.numero_competidor

    from flask_mail import Message

    texto = f"Hola {nombre},\n\nGracias por registrarte al Bombero Challenge.\nTu número de competidor es: #{nuevo_num}.\n\n¡Te esperamos!"

    html = f"""
    <html>
      <body style="font-family: Arial, sans-serif; padding: 20px; text-align: center;">
        <h2 style="color: #d32f2f;">Confirmación de Competencia</h2>
        <p>Hola <strong>{nombre}</strong>,</p>
        <p>Gracias por registrarte al <strong>Bombero Challenge</strong>.</p>
        <p><strong>Tu número de competidor es: #{nuevo_num}</strong></p>
        <p>¡Te esperamos!</p>
      </body>
    </html>
    """

    msg = Message(subject="Confirmación Bombero Challenge",
                  sender="cristian.rodriguez@bomberosdeleon.org",
                  recipients=[correo])

    msg.body = texto
    msg.html = html

    mail.send(msg)

    return render_template('confirmation_competencia.html', nombre=nombre, correo=correo, numero=nuevo_num)


@main.route('/submit_formulario', methods=['POST'])
def submit_formulario():
    try:
        preguntas = [
            'q1', 'q2', 'q3', 'q4', 'q5', 'q6', 'q7', 'q8', 
            'q9', 'q10', 'q11', 'q12', 'q13', 'q14', 'q15', 'q16', 
            'q17', 'q18', 'q19', 'q20', 'q21', 'q22', 'q23', 'q24', 
            'q25', 'q26', 'q27', 'q28', 'q29', 'q30', 'q31', 'q32', 
            'q33', 'q34', 'q35', 'q36', 'q37', 'q38', 'q39', 'q40', 
            'q41', 'q42', 'q43', 'q44', 'q45', 'q46', 'q47', 'q48', 
            'q49', 'q50', 'q51', 'q52', 'q53', 'q54', 'q55', 'q56', 
            'q57', 'q58', 'q59', 'q60', 'q61', 'q62', 'q63', 'q64'
        ]
        respuestas = {}
        for pregunta in preguntas:
            respuesta = request.form.get(pregunta)
            if respuesta:
                respuestas[pregunta] = int(respuesta) 
        nueva_respuesta = FormularioRespuesta(
            usuario_id=current_user.id,
            respuestas=json.dumps(respuestas),  
            fecha_creacion=datetime.now()
        )
        db.session.add(nueva_respuesta)
        db.session.commit()
        flash("¡Formulario enviado correctamente!", "success")
        return redirect(url_for('main.dashboard'))
    except Exception as e:
        db.session.rollback()
        flash(f"Error al enviar el formulario: {e}", "danger")
        return redirect(url_for('main.dashboard'))
@main.route('/add_evento', methods=['POST'])
@login_required
def add_evento():
    data = request.get_json()
    descripcion = data.get("descripcion")
    fecha = data.get("fecha")

    if descripcion and fecha:
        evento = Evento(
            descripcion=descripcion,
            fecha=datetime.strptime(fecha, "%Y-%m-%d").date(),
            user_id=current_user.id
        )
        db.session.add(evento)
        db.session.commit()
        return jsonify({"success": True, "evento_id": evento.id})

    return jsonify({"success": False}), 400

@main.route('/delete_evento/<int:evento_id>', methods=['DELETE'])
@login_required
def delete_evento(evento_id):
    evento = Evento.query.get_or_404(evento_id)
    db.session.delete(evento)
    db.session.commit()
    return jsonify({ "success": True })

@main.route('/admin/noticias', methods=['GET', 'POST'])
@login_required
def admin_noticias():
    if request.method == 'POST':
        titulo = request.form['titulo']
        descripcion = request.form['descripcion']
        imagen = request.files['imagen']
        link = request.form.get('link')

        if imagen:
            imagen.save(f"app/static/img/{imagen.filename}")

        nueva_noticia = Noticia(
            titulo=titulo,
            descripcion=descripcion,
            imagen=imagen.filename,
            link=link
        )
        db.session.add(nueva_noticia)
        db.session.commit()
        return redirect(url_for('main.admin_noticias'))

    noticias = Noticia.query.all()
    return render_template('admin_noticias.html', noticias=noticias)

@main.route('/admin/noticias/delete/<int:id>', methods=['GET', 'DELETE'])
@login_required
def delete_noticia(id):
    try:
        noticia = Noticia.query.get_or_404(id)
        db.session.delete(noticia)
        db.session.commit()
        if request.method == 'DELETE':
            return jsonify({'success': True, 'message': 'Noticia eliminada correctamente'})
        flash('Noticia eliminada correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        if request.method == 'DELETE':
            return jsonify({'success': False, 'message': str(e)}), 500
        flash(f'Error al eliminar noticia: {str(e)}', 'error')
    return redirect(url_for('main.admin_noticias'))

@main.route('/admin/permisos_evaluacion')
@login_required
def admin_permisos_evaluacion():
    """Página de admin para gestionar permisos de evaluación"""
    if current_user.username != 'admin':
        flash('Solo el admin puede acceder a esta página', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Obtener solo evaluadores permitidos (excluyendo BOMBERO ESPECIALIZADO)
    puestos_evaluadores = ['BOMBERO HABILITADO', 'SUBTENIENTE', 'TENIENTE', 'COORDINADOR']
    evaluadores = User.query.filter(
        User.puesto.in_(puestos_evaluadores),
        User.username != 'admin'
    ).all()
    
    return render_template('admin_permisos_evaluacion.html', evaluadores=evaluadores)

@main.route('/api/permisos_evaluador/<int:evaluador_id>')
@login_required
def api_permisos_evaluador(evaluador_id):
    """API que retorna los permisos actuales de un evaluador y usuarios disponibles"""
    if current_user.username != 'admin':
        return jsonify({'success': False, 'message': 'No autorizado'}), 403
    
    try:
        # Obtener permisos actuales
        permisos = PermisosEvaluacion.query.filter_by(evaluador_id=evaluador_id).all()
        permisos_dict = {p.evaluado_id: True for p in permisos}
        
        # Obtener todos los usuarios disponibles para evaluar (excepto el evaluador)
        usuarios = User.query.filter(User.id != evaluador_id, User.username != 'admin').all()
        
        usuarios_list = [
            {
                'id': u.id,
                'nombre': u.nombre,
                'puesto': u.puesto or 'N/A',
                'turno': u.turno or 'N/A',
                'username': u.username
            }
            for u in usuarios
        ]
        
        return jsonify({
            'success': True,
            'permisos': permisos_dict,
            'usuarios_disponibles': usuarios_list
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@main.route('/admin/guardar_permisos', methods=['POST'])
@login_required
def guardar_permisos():
    """Guardar los permisos de evaluación para un usuario"""
    if current_user.username != 'admin':
        return jsonify({'success': False, 'message': 'No autorizado'}), 403
    
    try:
        data = request.get_json()
        evaluador_id = data.get('evaluador_id')
        evaluados_ids = data.get('evaluados_ids', [])
        
        if not evaluador_id:
            return jsonify({'success': False, 'message': 'Evaluador no especificado'}), 400
        
        # Convertir a enteros y remover duplicados
        evaluados_ids = list(set([int(id) for id in evaluados_ids]))
        
        # Eliminar permisos anteriores
        PermisosEvaluacion.query.filter_by(evaluador_id=evaluador_id).delete()
        db.session.flush()
        
        # Agregar nuevos permisos
        for evaluado_id in evaluados_ids:
            try:
                permiso = PermisosEvaluacion(
                    evaluador_id=evaluador_id,
                    evaluado_id=evaluado_id
                )
                db.session.add(permiso)
            except Exception as inner_e:
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'message': f'Error al agregar permiso para usuario {evaluado_id}: {str(inner_e)}'
                }), 400
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Se guardaron {len(evaluados_ids)} permisos correctamente'
        })
    except ValueError as ve:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error de validación: {str(ve)}'}), 400
    except Exception as e:
        db.session.rollback()
        import traceback
        error_detail = traceback.format_exc()
        logging.error(f"Error en guardar_permisos: {error_detail}")
        return jsonify({
            'success': False,
            'message': f'Error al guardar permisos: {str(e)}'
        }), 500

@main.route('/delete_file/<filename>', methods=['DELETE'])
@login_required
def delete_file(filename):
    try:
        file_record = File.query.filter_by(filename=filename).first()
        if not file_record:
            return jsonify({'success': False, 'message': 'Archivo no encontrado'}), 404

        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(file_path):
            os.remove(file_path)

        db.session.delete(file_record)
        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@main.route('/list_files')
@login_required
def list_files():
    files = File.query.with_entities(File.filename).all()
    filenames = [f[0] for f in files]
    return jsonify({'files': filenames})

@main.route('/delete_aviso/<int:aviso_id>', methods=['POST'])
@login_required
def delete_aviso(aviso_id):
    if current_user.id != 2:
        return jsonify({'success': False, 'message': 'No tienes permiso para eliminar avisos.'}), 403
    
    aviso = Aviso.query.get_or_404(aviso_id)
    db.session.delete(aviso)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Aviso eliminado correctamente.'})


@main.route('/move_noticia/<int:id>/<string:direction>')
@login_required
def move_noticia(id, direction):
    noticia = Noticia.query.get_or_404(id)

    if direction == 'up':
        noticia_superior = Noticia.query.filter(Noticia.orden < noticia.orden).order_by(Noticia.orden.desc()).first()
        if noticia_superior:
            noticia.orden, noticia_superior.orden = noticia_superior.orden, noticia.orden

    elif direction == 'down':
        noticia_inferior = Noticia.query.filter(Noticia.orden > noticia.orden).order_by(Noticia.orden.asc()).first()
        if noticia_inferior:
            noticia.orden, noticia_inferior.orden = noticia_inferior.orden, noticia.orden

    db.session.commit()
    return redirect(url_for('main.admin_noticias'))

@main.route('/agregar_portal', methods=['POST'])
@login_required
def agregar_portal():
    if current_user.username.lower() == 'admin':
        nombre = (request.form.get('nombre') or '').strip()
        url = (request.form.get('url') or '').strip()
        if nombre and url:
            nuevo_portal = PortalWeb(nombre=nombre, url=url)
            restringido = request.form.get('visibilidad') == 'seleccionados'
            if restringido:
                ids = {int(value) for value in request.form.getlist('usuarios[]') if value.isdigit()}
                if not ids:
                    return jsonify({"success": False, "error": "Selecciona al menos un usuario"}), 400
                nuevo_portal.usuarios_permitidos = User.query.filter(
                    User.id.in_(ids),
                    ~db.func.lower(User.username).in_(USUARIOS_SIN_FUNCIONES)
                ).all()
            db.session.add(nuevo_portal)
            db.session.commit()
            return jsonify({"success": True})
        return jsonify({"success": False, "error": "Nombre y URL son obligatorios"}), 400
    return jsonify({"success": False, "error": "No autorizado"}), 403


@main.route('/eliminar_portal', methods=['POST'])
@login_required
def eliminar_portal():
    try:
        if current_user.username.lower() == 'admin':
            portal_id = request.json.get('id')
            portal = db.session.get(PortalWeb, portal_id)
            if portal:
                portal.usuarios_permitidos = []
                db.session.delete(portal)
                db.session.commit()
                return jsonify({"success": True})
        return jsonify({"success": False}), 403
    except Exception as e:
        logging.getLogger(__name__).exception("Error al eliminar portal")
        return jsonify({"success": False, "error": str(e)}), 500


@main.route('/guardar_contacto', methods=['POST'])
@login_required
def guardar_contacto():

    try:
        nombre = request.form['nombre_contacto']
        parentesco = request.form['parentesco']
        telefono = request.form['telefono_contacto']
        calle = request.form['calle_numero']
        colonia = request.form['colonia']
        otro = request.form.get('otro_parentesco')

        contacto_existente = ContactoEmergencia.query.filter_by(
            user_id=current_user.id
        ).first()

        if contacto_existente:
            # Actualiza si ya existe registro
            contacto_existente.username = current_user.username
            contacto_existente.nombre_contacto = nombre
            contacto_existente.parentesco = parentesco
            contacto_existente.telefono_contacto = telefono
            contacto_existente.calle_numero = calle
            contacto_existente.colonia = colonia
            contacto_existente.otro_parentesco = otro

            mensaje = "Contacto de emergencia actualizado correctamente"

        else:
            # Crea nuevo registro
            nuevo_contacto = ContactoEmergencia(
                user_id=current_user.id,
                username=current_user.username,
                nombre_contacto=nombre,
                parentesco=parentesco,
                telefono_contacto=telefono,
                calle_numero=calle,
                colonia=colonia,
                otro_parentesco=otro
            )

            db.session.add(nuevo_contacto)

            mensaje = "Contacto de emergencia registrado correctamente"

        db.session.commit()

        flash(mensaje, "success")

        return redirect(url_for('main.dashboard'))

    except Exception as e:
        db.session.rollback()
        flash(f"Error al guardar el contacto: {str(e)}", "danger")
        return redirect(url_for('main.dashboard'))


@main.route('/check_admin')
def check_admin():
    """Verifica si el usuario actual es administrador"""
    is_admin = current_user.is_authenticated and current_user.id == 2
    return jsonify({"is_admin": is_admin})


@main.route('/export_contactos_excel')
@login_required
def export_contactos_excel():
    """Exporta todos los contactos de emergencia a un archivo Excel"""
    try:
        # Verificar permisos de admin
        if current_user.id != 2:
            return jsonify({"error": "No tienes permisos para descargar este archivo"}), 403
        
        # Obtener todos los contactos
        contactos = ContactoEmergencia.query.all()
        
        # Preparar datos para Excel
        datos = []
        for contacto in contactos:
            usuario = User.query.get(contacto.user_id)
            
            # Determinar parentesco completo
            parentesco_completo = contacto.parentesco
            if contacto.parentesco == "Otro" and contacto.otro_parentesco:
                parentesco_completo = f"Otro ({contacto.otro_parentesco})"
            
            datos.append({
                'Nombre del Empleado': usuario.nombre if usuario else 'N/A',
                'Username': contacto.username,
                'Turno': usuario.turno if usuario else 'N/A',
                'Puesto': usuario.puesto if usuario else 'N/A',
                'Nombre del Contacto': contacto.nombre_contacto,
                'Parentesco': parentesco_completo,
                'Teléfono': contacto.telefono_contacto,
                'Calle y Número': contacto.calle_numero,
                'Colonia': contacto.colonia,
                'Fecha de Registro': contacto.fecha_registro.strftime('%Y-%m-%d %H:%M:%S') if contacto.fecha_registro else 'N/A'
            })
        
        # Crear DataFrame
        df = pd.DataFrame(datos)
        
        # Crear archivo Excel con formato
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Contactos', index=False)
            
            # Aplicar estilos
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = writer.sheets['Contactos']
            
            # Estilos
            header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar estilos al encabezado
            for col_num, value in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Aplicar bordes y ajustar ancho de columnas
            for col_num, col in enumerate(df.columns, 1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                max_length = 0
                
                for cell in worksheet[col_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[col_letter].width = min(adjusted_width, 50)
        
        output.seek(0)
        
        # Enviar archivo
        fecha = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"Contactos_Emergencia_{fecha}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logging.error(f"Error al exportar contactos a Excel: {str(e)}")
        return jsonify({"error": f"Error al descargar: {str(e)}"}), 500
